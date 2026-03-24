from __future__ import annotations

import base64
from copy import copy
from dataclasses import dataclass
from html import escape
from io import BytesIO
import json
import os
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
import requests
import streamlit as st

APP_TITLE = "SRMD MOM Generator"
DEFAULT_MODEL = "gpt-5.1"
MODEL_OPTIONS = ["gpt-5.1", "gpt-5-mini", "gpt-4.1"]
ZOHO_FALLBACK_OPTIONS = ["Vinay-Vivek", "NGH-A", "NGH-B", "NGH-C", "P2", "SRAH", "RU"]
NGH_PROJECT_OPTIONS = ["NGH-A", "NGH-B", "NGH-C"]
MANUAL_PROJECT_OPTION = "Other (Enter manually)"
ZOHO_PORTAL_ID_DEFAULT = "60062895348"
ZOHO_MOMS_MODULE_API_NAME_DEFAULT = "moms"  # api_name Zoho assigns to the MOMs custom module
DEFAULT_TEMPLATE_PATH = "mom_app/SRMD MOM Format.xlsx"
DISCUSSION_START_ROW = 19
BASE_DISCUSSION_ROWS = 9
FOOTER_START_ROW = 28
MIDDLE_DISCUSSION_TEMPLATE_ROW = 27
ATTENDEE_START_ROW = 10
ATTENDEE_ROW_COUNT = 6

CUSTOM_CSS = """
<style>
/* ── Fonts: Manrope (headlines) + Inter (body/labels) ── */
@import url('https://fonts.googleapis.com/css2?family=Manrope:wght@500;600;700;800&family=Inter:wght@300;400;500;600&display=swap');

/* ── Design tokens from DESIGN.md ── */
:root {
    --primary: #3E4C59;
    --primary-hover: #2d3a45;
    --secondary: #7A8C8C;
    --tertiary: #E0E4E8;
    --neutral: #121212;
    --bg: #FAFBFC;
    --surface: #FFFFFF;
    --surface-raised: #F4F6F8;
    --text-primary: #121212;
    --text-secondary: #3E4C59;
    --text-muted: #7A8C8C;
    --border: #E0E4E8;
    --border-hover: #C5CBD1;
    --radius: 8px;
    --radius-lg: 12px;
    --shadow-sm: 0 1px 3px rgba(18,18,18,0.06);
    --shadow-md: 0 4px 12px rgba(18,18,18,0.08);
    --transition: 0.18s ease;
}

html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    color: var(--text-primary);
}
h1, h2, h3, h4, h5, h6 {
    font-family: 'Manrope', 'Inter', sans-serif !important;
}

/* ── Light page background ── */
.stApp {
    background: var(--bg) !important;
}

/* ── Hide default chrome ── */
#MainMenu, footer {visibility: hidden;}
header[data-testid="stHeader"] {background: transparent !important;}

/* ── Branded header ── */
.app-header {
    background: var(--surface);
    padding: 1.6rem 2rem;
    border-radius: var(--radius-lg);
    margin-bottom: 1.5rem;
    border: 1px solid var(--border);
    box-shadow: var(--shadow-sm);
    display: flex;
    align-items: center;
    gap: 1rem;
}
.app-header-icon {
    width: 44px;
    height: 44px;
    background: var(--primary);
    border-radius: var(--radius);
    display: flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
}
.app-header-icon svg {
    width: 22px;
    height: 22px;
}
.app-header-text h1 {
    color: var(--neutral);
    font-size: 1.35rem;
    font-weight: 700;
    margin: 0;
    letter-spacing: -0.4px;
    line-height: 1.3;
}
.app-header-text p {
    color: var(--text-muted);
    font-family: 'Inter', sans-serif !important;
    font-size: 0.82rem;
    margin: 0.15rem 0 0 0;
    font-weight: 400;
}

/* ── Section headers ── */
.section-header {
    font-family: 'Manrope', sans-serif !important;
    font-size: 0.7rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: var(--text-muted);
    margin-bottom: 0.75rem;
    padding-bottom: 0.5rem;
    border-bottom: 1.5px solid var(--border);
}

/* ── Labels ── */
.stTextInput label, .stTextArea label, .stSelectbox label,
.stMultiSelect label, .stFileUploader label {
    font-family: 'Inter', sans-serif !important;
    font-size: 0.82rem !important;
    font-weight: 500 !important;
    color: var(--text-secondary) !important;
}

/* ── Inputs ── */
.stTextInput input, .stTextArea textarea {
    border-radius: var(--radius) !important;
    border: 1px solid var(--border) !important;
    background: var(--surface) !important;
    color: var(--text-primary) !important;
    font-size: 0.88rem !important;
    font-family: 'Inter', sans-serif !important;
    transition: border-color var(--transition), box-shadow var(--transition);
}
.stTextInput input::placeholder, .stTextArea textarea::placeholder {
    color: var(--text-muted) !important;
    opacity: 0.7;
}
.stTextInput input:focus, .stTextArea textarea:focus {
    border-color: var(--primary) !important;
    box-shadow: 0 0 0 3px rgba(62,76,89,0.1) !important;
}

/* ── Select boxes ── */
.stSelectbox [data-baseweb="select"] > div {
    border-radius: var(--radius) !important;
    border-color: var(--border) !important;
    background: var(--surface) !important;
    transition: border-color var(--transition);
}
.stSelectbox [data-baseweb="select"] > div:hover {
    border-color: var(--border-hover) !important;
}

/* ── All buttons — force readable text on their background ── */
button[data-testid="stBaseButton-primary"],
button[data-testid="stBaseButton-primary"] *,
button[data-testid="stBaseButton-primary"] p,
button[data-testid="stBaseButton-primary"] span {
    background: var(--primary) !important;
    border: none !important;
    border-radius: var(--radius) !important;
    padding: 0.6rem 1.8rem !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.88rem !important;
    letter-spacing: 0.2px !important;
    color: #ffffff !important;
    box-shadow: var(--shadow-sm) !important;
    transition: all var(--transition) !important;
}
button[data-testid="stBaseButton-primary"]:hover {
    background: var(--primary-hover) !important;
    box-shadow: var(--shadow-md) !important;
    transform: translateY(-1px) !important;
}

/* ── Secondary / download buttons ── */
.stDownloadButton > button,
button[data-testid="stBaseButton-secondary"] {
    border-radius: var(--radius) !important;
    border: 1px solid var(--border) !important;
    background: var(--surface) !important;
    color: var(--text-secondary) !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.84rem !important;
    box-shadow: var(--shadow-sm) !important;
    transition: all var(--transition) !important;
}
.stDownloadButton > button *,
button[data-testid="stBaseButton-secondary"] * {
    color: var(--text-secondary) !important;
}
.stDownloadButton > button:hover,
button[data-testid="stBaseButton-secondary"]:hover {
    border-color: var(--primary) !important;
    color: var(--primary) !important;
    background: var(--surface-raised) !important;
    box-shadow: var(--shadow-md) !important;
}
.stDownloadButton > button:hover *,
button[data-testid="stBaseButton-secondary"]:hover * {
    color: var(--primary) !important;
}

/* ── File uploader button & all other Streamlit buttons ── */
div[data-testid="stFileUploader"] button,
div[data-testid="stFileUploader"] button * {
    background: var(--surface) !important;
    color: var(--text-secondary) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--radius) !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.82rem !important;
}
div[data-testid="stFileUploader"] button:hover,
div[data-testid="stFileUploader"] button:hover * {
    border-color: var(--primary) !important;
    color: var(--primary) !important;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    gap: 0;
    border-bottom: 1.5px solid var(--border);
}
.stTabs [data-baseweb="tab"] {
    font-family: 'Inter', sans-serif;
    font-weight: 500;
    font-size: 0.84rem;
    padding: 0.6rem 1rem;
    color: var(--text-muted);
    border-bottom: 2px solid transparent;
    transition: all 0.15s ease;
}
.stTabs [aria-selected="true"] {
    color: var(--primary) !important;
    border-bottom-color: var(--primary) !important;
    font-weight: 600;
}

/* ── Alerts ── */
div[data-testid="stAlert"] {
    border-radius: var(--radius) !important;
    font-size: 0.84rem;
    border-left-width: 4px !important;
}

/* ── Expander ── */
.streamlit-expanderHeader {
    font-family: 'Inter', sans-serif !important;
    font-size: 0.84rem !important;
    font-weight: 500 !important;
    color: var(--text-secondary) !important;
}

/* ── Dataframe ── */
.stDataFrame {
    border-radius: var(--radius) !important;
    overflow: hidden;
    border: 1px solid var(--border);
    box-shadow: var(--shadow-sm);
}

/* ── Divider ── */
hr {
    border-color: var(--border) !important;
    margin: 1rem 0 !important;
}

/* ── File uploader ── */
div[data-testid="stFileUploader"] section {
    border-radius: var(--radius) !important;
    border: 1.5px dashed var(--border) !important;
    background: var(--surface-raised) !important;
    transition: border-color var(--transition);
}
div[data-testid="stFileUploader"] section:hover {
    border-color: var(--secondary) !important;
}

/* ── Sidebar ── */
section[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid var(--border) !important;
}
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stTextInput label {
    font-size: 0.72rem;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    color: var(--text-muted) !important;
}

/* ── Markdown text inside app ── */
[data-testid="stMarkdownContainer"] p {
    font-family: 'Inter', sans-serif !important;
    color: var(--text-primary);
}
[data-testid="stMarkdownContainer"] strong {
    font-weight: 600;
    color: var(--text-secondary);
}
</style>
"""

SYSTEM_PROMPT = """
You convert raw site-visit meeting notes into a structured Minutes of Meeting record.

Rules:
- Prefer any user-supplied field over inference.
- Infer missing details only from the notes and optional context.
- Keep wording professional, concise, and suitable for an Excel MOM register.
- Split the discussion into clear action-oriented points.
- Return strict JSON only.
- If a field is genuinely unclear, use a sensible placeholder instead of leaving it blank.
- Infer action status from language cues: "completed", "done", "closed", "approved" → "Closed";
  "pending", "yet to", "not done" → "Open"; "in progress", "partially done", "ongoing" → "In Progress";
  "deferred", "postponed", "on hold" → "Deferred". Default to "Open" if unclear.
- Extract responsible_party and target_date when the notes mention them (e.g. "XYZ to submit by 20 March"
  → responsible_party: "XYZ", target_date: "20 March"). Use empty strings when not mentioned.
- If the notes mention a next meeting date or place, populate next_meeting_date and next_meeting_place.
""".strip()


@dataclass
class DiscussionPoint:
    point_of_discussion: str
    discipline_of_work: str
    conclusion_or_remark: str
    responsible_party: str = ""
    target_date: str = ""
    status: str = "Open"


@dataclass
class MeetingRecord:
    project_name: str
    meeting_title: str
    meeting_date: str
    place: str
    attendees: list[str]
    discussion_points: list[DiscussionPoint]
    mom_number: str = ""
    next_meeting_date: str = ""
    next_meeting_place: str = ""


@dataclass
class ExistingMomContext:
    project_name: str = ""
    meeting_title: str = ""
    meeting_date: str = ""
    place: str = ""
    attendees: list[str] | None = None
    discussion_notes: str = ""
    workbook_text: str = ""


@dataclass
class AttendeeEntry:
    full_name: str
    agency: str = ""


def sanitize_text(value: Any, fallback: str) -> str:
    text = str(value or "").strip()
    return text or fallback


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return cleaned.strip("._") or "mom_report"


def get_api_key() -> str:
    secret_value = st.secrets.get("OPENAI_API_KEY", "")
    if secret_value:
        return str(secret_value)
    return os.getenv("OPENAI_API_KEY", "")


def _get_zoho_secret(key: str) -> str:
    return str(st.secrets.get(key, "") or os.getenv(key, ""))


def _get_zoho_access_token(client_id: str, client_secret: str, refresh_token: str) -> str:
    """Exchange refresh token for access token (India region). Raises on failure."""
    resp = requests.post(
        "https://accounts.zoho.in/oauth/v2/token",
        params={
            "grant_type": "refresh_token",
            "client_id": client_id,
            "client_secret": client_secret,
            "refresh_token": refresh_token,
        },
        timeout=10,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


@st.cache_data(ttl=300)
def _fetch_zoho_projects_raw() -> tuple[dict[str, str], str]:
    """Returns ({project_name: project_id}, error_msg) for active projects containing 'project'.
    error_msg is "" on success, otherwise describes what went wrong."""
    client_id = _get_zoho_secret("ZOHO_CLIENT_ID")
    client_secret = _get_zoho_secret("ZOHO_CLIENT_SECRET")
    refresh_token = _get_zoho_secret("ZOHO_REFRESH_TOKEN")
    portal_id = _get_zoho_secret("ZOHO_PORTAL_ID") or ZOHO_PORTAL_ID_DEFAULT
    if not (client_id and client_secret and refresh_token):
        return {}, "Zoho credentials not configured."
    try:
        access_token = _get_zoho_access_token(client_id, client_secret, refresh_token)
        projects: list[dict] = []
        page = 1
        while True:
            resp = requests.get(
                f"https://projectsapi.zoho.in/restapi/portal/{portal_id}/projects/",
                headers={"Authorization": f"Zoho-oauthtoken {access_token}"},
                params={"status": "active", "per_page": 100, "page": page},
                timeout=10,
            )
            if not resp.ok:
                return {}, f"Zoho API error ({resp.status_code}): {resp.text[:200]}"
            batch = resp.json().get("projects", [])
            if not batch:
                break
            projects.extend(batch)
            if len(batch) < 100:
                break
            page += 1
        mapping = {
            p["name"].strip(): str(p["id"])
            for p in projects
            if "project" in p["name"].lower()
        }
        return mapping, ""
    except Exception as exc:
        return {}, f"Zoho projects fetch failed: {exc}"


@st.cache_data(ttl=300)
def fetch_zoho_project_options() -> tuple[list[str], str]:
    """Returns (project_names, error_msg). Falls back to ZOHO_FALLBACK_OPTIONS on error."""
    mapping, err = _fetch_zoho_projects_raw()
    if mapping:
        return sorted(mapping.keys()), ""
    return ZOHO_FALLBACK_OPTIONS, err


def get_zoho_project_id(name: str) -> str | None:
    """Returns the Zoho project ID for a given project name, or None if not found."""
    mapping, _ = _fetch_zoho_projects_raw()
    return mapping.get(name)


@st.cache_data(ttl=3600)
def _get_moms_module_api_name(portal_id: str) -> str:
    """Returns the api_name for the 'MOMs' custom module. Cached 1 hour.

    Resolution order:
      1. ZOHO_MOMS_MODULE_NAME secret (manual override)
      2. GET /api/v3/portal/{portal_id}/settings/modules (requires ZohoProjects.custom_fields.READ)
      3. Hardcoded default "moms"
    """
    manual = _get_zoho_secret("ZOHO_MOMS_MODULE_NAME")
    if manual:
        return manual

    client_id = _get_zoho_secret("ZOHO_CLIENT_ID")
    client_secret = _get_zoho_secret("ZOHO_CLIENT_SECRET")
    refresh_token = _get_zoho_secret("ZOHO_REFRESH_TOKEN")

    if client_id and client_secret and refresh_token:
        try:
            access_token = _get_zoho_access_token(client_id, client_secret, refresh_token)
            resp = requests.get(
                f"https://projectsapi.zoho.in/api/v3/portal/{portal_id}/settings/modules",
                headers={"Authorization": f"Zoho-oauthtoken {access_token}"},
                timeout=10,
            )
            if resp.ok:
                for mod in resp.json().get("modules", []):
                    if mod.get("singular_name", "").strip().lower() == "mom" or \
                       mod.get("plural_name", "").strip().lower() == "moms" or \
                       mod.get("api_name", "").strip().lower() == "moms":
                        return str(mod["api_name"])
        except Exception:
            pass

    return ZOHO_MOMS_MODULE_API_NAME_DEFAULT


def clean_lines(value: str) -> list[str]:
    return [line.strip(" -\t") for line in value.splitlines() if line.strip()]


def parse_json_response(raw_text: str) -> dict[str, Any]:
    raw_text = raw_text.strip()
    if not raw_text:
        raise ValueError("The model returned an empty response.")

    try:
        return json.loads(raw_text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", raw_text, re.DOTALL)
        if not match:
            raise ValueError("The model response was not valid JSON.")
        return json.loads(match.group(0))


def unique_nonempty(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = value.strip()
        if not cleaned:
            continue
        lowered = cleaned.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        result.append(cleaned)
    return result


def extract_attendees_from_text(*texts: str) -> list[str]:
    attendees: list[str] = []
    section_headers = ("attendees", "attendee", "present", "participants", "present members")
    stop_headers = ("discussion", "agenda", "points", "summary", "remarks", "conclusion", "action")
    attendee_markers = ("mr", "mrs", "ms", "dr", "er", "architect", "consultant", "client", "contractor", "engineer")

    for text in texts:
        if not text:
            continue
        lines = [line.strip() for line in text.splitlines()]
        capture = False
        for line in lines:
            if not line:
                if capture:
                    capture = False
                continue
            normalized = line.strip(" :-").lower()
            if any(normalized.startswith(header) for header in section_headers):
                capture = True
                extracted = line.split(":", 1)[1].strip() if ":" in line else ""
                if extracted:
                    attendees.extend(re.split(r",|;|\n", extracted))
                continue
            stripped_numbered_line = re.sub(r"^\d+[\).:-]\s*", "", line).strip()
            numbered_attendee_line = (
                re.match(r"^\d+[\).:-]", line)
                and (
                    " - " in stripped_numbered_line
                    or any(marker in stripped_numbered_line.lower() for marker in attendee_markers)
                )
            )
            if capture and (
                any(normalized.startswith(header) for header in stop_headers)
                or (re.match(r"^\d+[\).:-]", line) and not numbered_attendee_line)
            ):
                capture = False
            if capture:
                attendees.extend(re.split(r",|;", stripped_numbered_line if numbered_attendee_line else line))

    cleaned: list[str] = []
    for attendee in attendees:
        attendee = re.sub(r"^\d+[\).:-]?\s*", "", attendee).strip(" -")
        if len(attendee) < 3:
            continue
        if any(token in attendee.lower() for token in stop_headers):
            continue
        cleaned.append(attendee)
    return unique_nonempty(cleaned)


def parse_attendee_entries(attendees: list[str]) -> list[AttendeeEntry]:
    parsed: list[AttendeeEntry] = []
    for attendee in attendees:
        text = attendee.strip()
        if not text:
            continue
        full_name = text
        agency = ""
        if " - " in text:
            full_name, agency = [part.strip() for part in text.split(" - ", 1)]
        elif "," in text:
            full_name, agency = [part.strip() for part in text.split(",", 1)]
        parsed.append(AttendeeEntry(full_name=full_name or text, agency=agency))
    return parsed


def set_project_state(project_name: str) -> None:
    project_name = project_name.strip()
    parsed_gh_projects = [option for option in NGH_PROJECT_OPTIONS if option.lower() in project_name.lower()]
    if parsed_gh_projects:
        st.session_state["project_name_select"] = parsed_gh_projects[0]
        st.session_state["project_name_ngh_multi"] = parsed_gh_projects
        st.session_state["project_name_custom_input"] = ""
    elif project_name in PROJECT_OPTIONS:
        st.session_state["project_name_select"] = project_name
        st.session_state["project_name_ngh_multi"] = []
        st.session_state["project_name_custom_input"] = ""
    elif project_name:
        st.session_state["project_name_select"] = MANUAL_PROJECT_OPTION
        st.session_state["project_name_ngh_multi"] = []
        st.session_state["project_name_custom_input"] = project_name


def split_date_place(value: str) -> tuple[str, str]:
    text = value.strip()
    if not text:
        return "", ""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return "", ""
    if len(lines) == 1 and "place:" in lines[0].lower():
        date_part, place_part = re.split(r"place\s*:", lines[0], maxsplit=1, flags=re.IGNORECASE)
        return date_part.strip(" :-"), place_part.strip()

    date_part = lines[0]
    place_part = ""
    for line in lines[1:]:
        if line.lower().startswith("place:"):
            place_part = line.split(":", 1)[1].strip()
        elif not place_part:
            place_part = line
    return date_part, place_part


def extract_existing_mom_context(uploaded_file: Any) -> ExistingMomContext:
    workbook = load_workbook(BytesIO(uploaded_file.getvalue()), data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    workbook_values: list[str] = []
    for row in worksheet.iter_rows(values_only=True):
        for value in row:
            text = str(value).strip() if value is not None else ""
            if text:
                workbook_values.append(text)

    discussion_lines: list[str] = []
    for row in range(DISCUSSION_START_ROW, worksheet.max_row + 1):
        point = str(worksheet[f"C{row}"].value or "").strip()
        discipline = str(worksheet[f"E{row}"].value or "").strip()
        remark = str(worksheet[f"F{row}"].value or "").strip()
        if point:
            line = point
            if discipline:
                line += f" | Discipline: {discipline}"
            if remark:
                line += f" | Remark: {remark}"
            discussion_lines.append(line)

    attendee_candidates: list[str] = []
    for row in range(ATTENDEE_START_ROW, ATTENDEE_START_ROW + ATTENDEE_ROW_COUNT):
        full_name = str(worksheet[f"C{row}"].value or "").strip()
        agency = str(worksheet[f"D{row}"].value or "").strip()
        if full_name:
            attendee_candidates.append(f"{full_name} - {agency}".strip(" -"))
    attendee_candidates = unique_nonempty(attendee_candidates)

    date_value, place_value = split_date_place(str(worksheet["D5"].value or ""))
    discussion_notes = "\n".join(discussion_lines)
    workbook_text = "\n".join(unique_nonempty(workbook_values))

    return ExistingMomContext(
        project_name=str(worksheet["B2"].value or "").strip(),
        meeting_title=str(worksheet["D4"].value or "").strip(),
        meeting_date=date_value,
        place=place_value,
        attendees=attendee_candidates,
        discussion_notes=discussion_notes,
        workbook_text=workbook_text,
    )


_VALID_STATUSES = {"Open", "In Progress", "Closed", "Deferred"}


def _validate_status(value: str) -> str:
    return value if value in _VALID_STATUSES else "Open"


def infer_discipline(text: str) -> str:
    lowered = text.lower()
    mapping = {
        "Civil": ["concrete", "slab", "column", "beam", "plaster", "brick", "masonry", "floor", "excavation"],
        "MEP": ["electrical", "wiring", "cable", "lighting", "plumbing", "drain", "pipe", "hvac", "fire fighting"],
        "Architecture": ["finish", "façade", "facade", "paint", "tile", "door", "window", "ceiling", "elevation"],
        "Safety": ["safety", "ppe", "barricade", "housekeeping", "hazard", "incident"],
        "Planning": ["schedule", "timeline", "delay", "handover", "approval", "submission", "procurement"],
    }
    for discipline, keywords in mapping.items():
        if any(keyword in lowered for keyword in keywords):
            return discipline
    return "General"


def heuristic_discussion_points(notes: str) -> list[DiscussionPoint]:
    lines = clean_lines(notes)
    if not lines:
        lines = [segment.strip() for segment in re.split(r"(?<=[.!?])\s+", notes) if segment.strip()]

    points: list[DiscussionPoint] = []
    for line in lines:
        normalized = re.sub(r"^\d+[\).:-]?\s*", "", line).strip()
        if not normalized:
            continue

        lowered = normalized.lower()
        conclusion = "Review on site and close the action as discussed."
        status = "Open"
        if any(token in lowered for token in ["approved", "completed", "closed", "done"]):
            conclusion = "Noted as completed / accepted during the meeting."
            status = "Closed"
        elif any(token in lowered for token in ["in progress", "partially", "ongoing"]):
            status = "In Progress"
        elif any(token in lowered for token in ["deferred", "postponed", "on hold"]):
            status = "Deferred"
        elif any(token in lowered for token in ["provide", "submit", "share", "revise", "rectify"]):
            conclusion = "Concerned team to take action and update in the next review."

        party_match = re.search(r"\b([A-Z][a-z]+(?: [A-Z][a-z]+)*)\s+to\s+\w", normalized)
        responsible_party = party_match.group(1) if party_match else ""

        date_match = re.search(r"\bby\s+(\d{1,2}\s+\w+(?:\s+\d{4})?)", normalized, re.IGNORECASE)
        target_date = date_match.group(1) if date_match else ""

        points.append(
            DiscussionPoint(
                point_of_discussion=normalized,
                discipline_of_work=infer_discipline(normalized),
                conclusion_or_remark=conclusion,
                responsible_party=responsible_party,
                target_date=target_date,
                status=status,
            )
        )

    return points[:25] or [
        DiscussionPoint(
            point_of_discussion="Review the submitted site visit notes and confirm the action items.",
            discipline_of_work="General",
            conclusion_or_remark="Action owners to align and issue an updated status after the meeting.",
        )
    ]


def heuristic_meeting_record(
    project_name: str,
    meeting_title: str,
    meeting_date: str,
    place: str,
    attendees_text: str,
    notes: str,
    mom_number: str = "",
    next_meeting_date: str = "",
    next_meeting_place: str = "",
) -> MeetingRecord:
    attendees = unique_nonempty(clean_lines(attendees_text) + extract_attendees_from_text(attendees_text, notes))
    return MeetingRecord(
        project_name=sanitize_text(project_name, "PROJECT NAME TO BE CONFIRMED"),
        meeting_title=sanitize_text(meeting_title, "Site Visit Meeting"),
        meeting_date=sanitize_text(meeting_date, "Date to be confirmed"),
        place=sanitize_text(place, "Place to be confirmed"),
        attendees=attendees or ["Attendee details to be confirmed from the meeting notes."],
        discussion_points=heuristic_discussion_points(notes),
        mom_number=mom_number.strip(),
        next_meeting_date=next_meeting_date.strip(),
        next_meeting_place=next_meeting_place.strip(),
    )


def build_prompt(
    project_name: str,
    meeting_title: str,
    meeting_date: str,
    place: str,
    attendees_text: str,
    notes: str,
    extra_context: str,
    mom_number: str = "",
    next_meeting_date: str = "",
    next_meeting_place: str = "",
) -> str:
    payload = {
        "user_inputs": {
            "project_name": project_name.strip(),
            "meeting_title": meeting_title.strip(),
            "meeting_date": meeting_date.strip(),
            "place": place.strip(),
            "attendees": clean_lines(attendees_text),
            "mom_number": mom_number.strip(),
            "next_meeting_date": next_meeting_date.strip(),
            "next_meeting_place": next_meeting_place.strip(),
        },
        "extra_context": extra_context.strip(),
        "meeting_notes": notes.strip(),
        "output_schema": {
            "project_name": "string",
            "meeting_title": "string",
            "meeting_date": "string",
            "place": "string",
            "attendees": ["short attendee strings"],
            "mom_number": "use the user-supplied value as-is",
            "next_meeting_date": "next meeting date if mentioned in notes, else empty string",
            "next_meeting_place": "next meeting place if mentioned in notes, else empty string",
            "discussion_points": [
                {
                    "point_of_discussion": "string under 180 characters",
                    "discipline_of_work": "short label like Civil / MEP / Architecture / Safety / Planning / General",
                    "conclusion_or_remark": "string under 160 characters",
                    "responsible_party": "name or team responsible, empty string if not mentioned",
                    "target_date": "target completion date, empty string if not mentioned",
                    "status": "Open | In Progress | Closed | Deferred",
                }
            ],
        },
    }
    return json.dumps(payload, indent=2)


def normalize_meeting_record(payload: dict[str, Any]) -> MeetingRecord:
    raw_points = payload.get("discussion_points")
    points: list[DiscussionPoint] = []
    if isinstance(raw_points, list):
        for item in raw_points:
            if not isinstance(item, dict):
                continue
            points.append(
                DiscussionPoint(
                    point_of_discussion=sanitize_text(
                        item.get("point_of_discussion"),
                        "Discussion point to be confirmed.",
                    ),
                    discipline_of_work=sanitize_text(item.get("discipline_of_work"), "General"),
                    conclusion_or_remark=sanitize_text(
                        item.get("conclusion_or_remark"),
                        "Action owner to review and update.",
                    ),
                    responsible_party=str(item.get("responsible_party") or "").strip(),
                    target_date=str(item.get("target_date") or "").strip(),
                    status=_validate_status(str(item.get("status") or "Open").strip()),
                )
            )

    attendees_value = payload.get("attendees")
    attendees = []
    if isinstance(attendees_value, list):
        attendees = [str(item).strip() for item in attendees_value if str(item).strip()]

    return MeetingRecord(
        project_name=sanitize_text(payload.get("project_name"), "PROJECT NAME TO BE CONFIRMED"),
        meeting_title=sanitize_text(payload.get("meeting_title"), "Site Visit Meeting"),
        meeting_date=sanitize_text(payload.get("meeting_date"), "Date to be confirmed"),
        place=sanitize_text(payload.get("place"), "Place to be confirmed"),
        attendees=attendees or ["Attendee details to be confirmed from the meeting notes."],
        discussion_points=points
        or [
            DiscussionPoint(
                point_of_discussion="Review the submitted site visit notes and confirm the action items.",
                discipline_of_work="General",
                conclusion_or_remark="Action owners to align and issue an updated status after the meeting.",
            )
        ],
        mom_number=str(payload.get("mom_number") or "").strip(),
        next_meeting_date=str(payload.get("next_meeting_date") or "").strip(),
        next_meeting_place=str(payload.get("next_meeting_place") or "").strip(),
    )


def generate_meeting_record(
    api_key: str,
    model: str,
    project_name: str,
    meeting_title: str,
    meeting_date: str,
    place: str,
    attendees_text: str,
    notes: str,
    extra_context: str,
    mom_number: str = "",
    next_meeting_date: str = "",
    next_meeting_place: str = "",
) -> MeetingRecord:
    fallback = heuristic_meeting_record(
        project_name=project_name,
        meeting_title=meeting_title,
        meeting_date=meeting_date,
        place=place,
        attendees_text=attendees_text,
        notes=notes,
        mom_number=mom_number,
        next_meeting_date=next_meeting_date,
        next_meeting_place=next_meeting_place,
    )
    if not api_key:
        return fallback

    client = OpenAI(api_key=api_key)
    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "system",
                "content": [{"type": "input_text", "text": SYSTEM_PROMPT}],
            },
            {
                "role": "user",
                "content": [{"type": "input_text", "text": build_prompt(
                    project_name, meeting_title, meeting_date, place, attendees_text,
                    notes, extra_context, mom_number, next_meeting_date, next_meeting_place,
                )}],
            },
        ],
    )
    record = normalize_meeting_record(parse_json_response(response.output_text))
    # User-supplied values always override AI inference
    if mom_number:
        record.mom_number = mom_number
    if next_meeting_date:
        record.next_meeting_date = next_meeting_date
    if next_meeting_place:
        record.next_meeting_place = next_meeting_place
    return record


def template_bytes() -> bytes:
    with open(DEFAULT_TEMPLATE_PATH, "rb") as template_file:
        return template_file.read()


def copy_row_style(ws: Any, source_row: int, target_row: int, start_col: int = 1, end_col: int = 10) -> None:
    source_height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].height = source_height
    for col in range(start_col, end_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if isinstance(source, MergedCell):
            continue
        target._style = copy(source._style)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)


def rebuild_dynamic_layout(ws: Any, extra_rows: int) -> None:
    if extra_rows <= 0:
        return

    saved_heights = {row: ws.row_dimensions[row].height for row in range(FOOTER_START_ROW, 35)}

    ws.move_range(f"B{FOOTER_START_ROW}:F34", rows=extra_rows, cols=0, translate=True)

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= FOOTER_START_ROW:
            ws.merged_cells.ranges.remove(merged_range)

    for row in range(FOOTER_START_ROW, FOOTER_START_ROW + extra_rows):
        copy_row_style(ws, MIDDLE_DISCUSSION_TEMPLATE_ROW, row, start_col=2, end_col=6)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

    for row in range(FOOTER_START_ROW + extra_rows, 35 + extra_rows):
        original_row = row - extra_rows
        if original_row in saved_heights:
            ws.row_dimensions[row].height = saved_heights[original_row]

    ws.merge_cells(start_row=28 + extra_rows, start_column=3, end_row=28 + extra_rows, end_column=4)
    ws.merge_cells(start_row=29 + extra_rows, start_column=2, end_row=29 + extra_rows, end_column=6)
    ws.merge_cells(start_row=30 + extra_rows, start_column=2, end_row=30 + extra_rows, end_column=6)
    ws.merge_cells(start_row=31 + extra_rows, start_column=2, end_row=33 + extra_rows, end_column=6)
    ws.merge_cells(start_row=34 + extra_rows, start_column=3, end_row=34 + extra_rows, end_column=4)


def fill_attendees(ws: Any, attendees: list[str]) -> None:
    parsed_attendees = parse_attendee_entries(attendees)
    for index, attendee in enumerate(parsed_attendees[:ATTENDEE_ROW_COUNT], start=ATTENDEE_START_ROW):
        ws[f"B{index}"] = index - ATTENDEE_START_ROW + 1
        ws[f"C{index}"] = attendee.full_name
        ws[f"D{index}"] = attendee.agency
        ws[f"C{index}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{index}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if len(parsed_attendees) > ATTENDEE_ROW_COUNT:
        overflow_items = parsed_attendees[ATTENDEE_ROW_COUNT:]
        overflow_names = ", ".join(item.full_name for item in overflow_items)
        overflow_agencies = ", ".join(item.agency for item in overflow_items if item.agency)
        ws[f"C{ATTENDEE_START_ROW + ATTENDEE_ROW_COUNT - 1}"] = (
            f"{ws[f'C{ATTENDEE_START_ROW + ATTENDEE_ROW_COUNT - 1}'].value}\nAdditional: {overflow_names}"
        )
        if overflow_agencies:
            ws[f"D{ATTENDEE_START_ROW + ATTENDEE_ROW_COUNT - 1}"] = (
                f"{ws[f'D{ATTENDEE_START_ROW + ATTENDEE_ROW_COUNT - 1}'].value or ''}\n{overflow_agencies}".strip()
            )


def build_conclusion_cell_text(point: DiscussionPoint) -> str:
    parts: list[str] = []
    if point.status and point.status != "Open":
        parts.append(f"[{point.status}]")
    parts.append(point.conclusion_or_remark)
    suffix_parts: list[str] = []
    if point.responsible_party:
        suffix_parts.append(f"Owner: {point.responsible_party}")
    if point.target_date:
        suffix_parts.append(f"Due: {point.target_date}")
    if suffix_parts:
        parts.append(" | ".join(suffix_parts))
    return " ".join(parts).strip()


def fill_discussion_table(ws: Any, discussion_points: list[DiscussionPoint]) -> None:
    points = discussion_points[:25]
    extra_rows = max(0, len(points) - BASE_DISCUSSION_ROWS)
    rebuild_dynamic_layout(ws, extra_rows)

    for index, point in enumerate(points, start=DISCUSSION_START_ROW):
        serial_no = index - DISCUSSION_START_ROW + 1
        ws[f"B{index}"] = serial_no
        ws[f"C{index}"] = point.point_of_discussion
        ws[f"E{index}"] = point.discipline_of_work
        ws[f"F{index}"] = build_conclusion_cell_text(point)

        for coord in [f"C{index}", f"E{index}", f"F{index}"]:
            ws[coord].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def estimate_line_count(text: str, width_units: float) -> int:
    if not text:
        return 1
    normalized = str(text).replace("\r", "")
    explicit_lines = normalized.split("\n")
    chars_per_line = max(12, int(width_units * 1.4))
    total_lines = 0
    for line in explicit_lines:
        total_lines += max(1, (len(line) // chars_per_line) + (1 if len(line) % chars_per_line else 0))
    return max(1, total_lines)


def adjust_sheet_layout(ws: Any, meeting_record: MeetingRecord) -> None:
    point_lengths = [len(point.point_of_discussion) for point in meeting_record.discussion_points] or [0]
    discipline_lengths = [len(point.discipline_of_work) for point in meeting_record.discussion_points] or [0]
    remark_lengths = [len(point.conclusion_or_remark) for point in meeting_record.discussion_points] or [0]

    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 39, min(48, 30 + max(point_lengths) / 8))
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width or 30, min(36, 24 + max(point_lengths) / 10))
    ws.column_dimensions["E"].width = max(ws.column_dimensions["E"].width or 13, min(20, 11 + max(discipline_lengths) / 5))
    ws.column_dimensions["F"].width = max(ws.column_dimensions["F"].width or 52, min(62, 32 + max(remark_lengths) / 5))

    for row in range(ATTENDEE_START_ROW, ATTENDEE_START_ROW + ATTENDEE_ROW_COUNT):
        name_lines = estimate_line_count(str(ws[f"C{row}"].value or ""), ws.column_dimensions["C"].width or 39)
        agency_lines = estimate_line_count(str(ws[f"D{row}"].value or ""), (ws.column_dimensions["D"].width or 30) + (ws.column_dimensions["F"].width or 52))
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 18, 16 * max(name_lines, agency_lines))

    for row in range(DISCUSSION_START_ROW, DISCUSSION_START_ROW + len(meeting_record.discussion_points)):
        point_lines = estimate_line_count(str(ws[f"C{row}"].value or ""), (ws.column_dimensions["C"].width or 40) + (ws.column_dimensions["D"].width or 30))
        discipline_lines = estimate_line_count(str(ws[f"E{row}"].value or ""), ws.column_dimensions["E"].width or 13)
        remark_lines = estimate_line_count(str(ws[f"F{row}"].value or ""), ws.column_dimensions["F"].width or 52)
        row_lines = max(point_lines, discipline_lines, remark_lines)
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 18, 15 * row_lines)


def build_workbook(template_content: bytes, meeting_record: MeetingRecord) -> bytes:
    workbook = load_workbook(BytesIO(template_content))
    worksheet = workbook[workbook.sheetnames[0]]

    worksheet["B2"] = meeting_record.project_name.upper()
    title_value = meeting_record.meeting_title
    if meeting_record.mom_number:
        title_value = f"{meeting_record.meeting_title} | Ref: {meeting_record.mom_number}"
    worksheet["D4"] = title_value
    worksheet["D5"] = f"{meeting_record.meeting_date}\nPlace: {meeting_record.place}"
    worksheet["D5"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    fill_attendees(worksheet, meeting_record.attendees)
    fill_discussion_table(worksheet, meeting_record.discussion_points)
    adjust_sheet_layout(worksheet, meeting_record)

    if meeting_record.next_meeting_date or meeting_record.next_meeting_place:
        extra_rows = max(0, len(meeting_record.discussion_points[:25]) - BASE_DISCUSSION_ROWS)
        footer_row = 29 + extra_rows
        next_parts = ["Next Meeting:"]
        if meeting_record.next_meeting_date:
            next_parts.append(f"Date: {meeting_record.next_meeting_date}")
        if meeting_record.next_meeting_place:
            next_parts.append(f"Place: {meeting_record.next_meeting_place}")
        worksheet[f"B{footer_row}"] = "  |  ".join(next_parts)
        worksheet[f"B{footer_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


_PDF_STATUS_COLORS: dict[str, str] = {
    "Open": "#ffe8e8",
    "In Progress": "#fff8e1",
    "Closed": "#e8f5e9",
    "Deferred": "#eeeeee",
}


def build_pdf_report(meeting_record: MeetingRecord) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("MOMTitle", parent=styles["Title"], textColor=colors.HexColor("#1f3a5f"))
    body_style = ParagraphStyle("MOMBody", parent=styles["BodyText"], leading=14, spaceAfter=5)
    small_style = ParagraphStyle("MOMSmall", parent=body_style, fontSize=9, leading=11)
    small_muted_style = ParagraphStyle("MOMMuted", parent=small_style, textColor=colors.HexColor("#555555"))

    story: list[Any] = [
        Paragraph("Minutes of Meeting", title_style),
        Spacer(1, 4),
        Paragraph(f"<b>Project:</b> {escape(meeting_record.project_name)}", body_style),
        Paragraph(f"<b>Meeting:</b> {escape(meeting_record.meeting_title)}", body_style),
        Paragraph(f"<b>Date:</b> {escape(meeting_record.meeting_date)}", body_style),
        Paragraph(f"<b>Place:</b> {escape(meeting_record.place)}", body_style),
    ]
    if meeting_record.mom_number:
        story.append(Paragraph(f"<b>MOM Ref:</b> {escape(meeting_record.mom_number)}", body_style))
    story += [
        Spacer(1, 4),
        Paragraph("<b>Attendees</b>", body_style),
    ]

    for attendee in meeting_record.attendees:
        story.append(Paragraph(f"- {escape(attendee)}", body_style))

    story.append(Spacer(1, 6))

    # 5-column table: Sr | Point | Discipline | Status | Conclusion/Remark
    table_data: list[Any] = [["Sr.", "Point of Discussion", "Discipline", "Status", "Conclusion / Remark"]]
    for index, point in enumerate(meeting_record.discussion_points, start=1):
        conclusion_lines = [escape(point.conclusion_or_remark)]
        if point.responsible_party or point.target_date:
            owner_parts: list[str] = []
            if point.responsible_party:
                owner_parts.append(f"Owner: {escape(point.responsible_party)}")
            if point.target_date:
                owner_parts.append(f"Due: {escape(point.target_date)}")
            conclusion_lines.append(f'<font size="8" color="#555555">{" | ".join(owner_parts)}</font>')
        table_data.append(
            [
                str(index),
                Paragraph(escape(point.point_of_discussion), small_style),
                Paragraph(escape(point.discipline_of_work), small_style),
                Paragraph(escape(point.status), small_style),
                Paragraph("<br/>".join(conclusion_lines), small_style),
            ]
        )

    table = Table(table_data, colWidths=[12 * mm, 68 * mm, 25 * mm, 18 * mm, 57 * mm], repeatRows=1)

    table_style_cmds: list[Any] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9ead3")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#8a8a8a")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]
    for row_idx, point in enumerate(meeting_record.discussion_points, start=1):
        bg_hex = _PDF_STATUS_COLORS.get(point.status, "#ffffff")
        table_style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor(bg_hex)))

    table.setStyle(TableStyle(table_style_cmds))
    story.append(table)

    if meeting_record.next_meeting_date or meeting_record.next_meeting_place:
        story.append(Spacer(1, 10))
        story.append(Paragraph("<b>Next Meeting</b>", body_style))
        if meeting_record.next_meeting_date:
            story.append(Paragraph(f"<b>Date:</b> {escape(meeting_record.next_meeting_date)}", body_style))
        if meeting_record.next_meeting_place:
            story.append(Paragraph(f"<b>Place:</b> {escape(meeting_record.next_meeting_place)}", body_style))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def build_email_draft(record: MeetingRecord) -> str:
    ref_suffix = f" ({record.mom_number})" if record.mom_number else ""
    subject = f"Subject: Minutes of Meeting – {record.project_name} | {record.meeting_date}{ref_suffix}"

    lines: list[str] = [
        subject,
        "",
        "Dear All,",
        "",
        (
            f"Please find below the Minutes of Meeting for the site visit held on {record.meeting_date}"
            + (f" at {record.place}" if record.place else "") + "."
        ),
        "",
        "--- MEETING DETAILS ---",
        f"Project     : {record.project_name}",
        f"Meeting     : {record.meeting_title}{ref_suffix}",
        f"Date        : {record.meeting_date}",
        f"Place       : {record.place}",
        "",
        "--- ATTENDEES ---",
    ]
    for attendee in record.attendees:
        lines.append(f"  • {attendee}")

    open_statuses = {"Open", "In Progress"}
    action_points = [
        (i, p) for i, p in enumerate(record.discussion_points, 1)
        if p.status in open_statuses or p.responsible_party
    ]
    lines += ["", "--- ACTION ITEMS ---"]
    if action_points:
        for i, point in action_points:
            owner = f" | Owner: {point.responsible_party}" if point.responsible_party else ""
            due = f" | Due: {point.target_date}" if point.target_date else ""
            status_tag = f" [{point.status}]" if point.status != "Open" else ""
            lines.append(f"{i}. {point.point_of_discussion}{status_tag}{owner}{due}")
    else:
        lines.append("  No open action items recorded.")

    lines += ["", "--- FULL DISCUSSION POINTS ---"]
    for i, point in enumerate(record.discussion_points, 1):
        lines.append(f"{i}. [{point.discipline_of_work}] {point.point_of_discussion}")
        lines.append(f"   Remark: {point.conclusion_or_remark}")
        detail_parts: list[str] = []
        if point.responsible_party:
            detail_parts.append(f"Owner: {point.responsible_party}")
        if point.target_date:
            detail_parts.append(f"Due: {point.target_date}")
        if detail_parts:
            lines.append(f"   {' | '.join(detail_parts)}")
        lines.append("")

    if record.next_meeting_date or record.next_meeting_place:
        lines += ["--- NEXT MEETING ---"]
        if record.next_meeting_date:
            lines.append(f"Date  : {record.next_meeting_date}")
        if record.next_meeting_place:
            lines.append(f"Place : {record.next_meeting_place}")
        lines.append("")

    lines += [
        "Please review and revert with any corrections within 48 hours.",
        "",
        "Regards,",
        "[Your Name]",
        "[Your Designation]",
        "[Organization]",
    ]
    return "\n".join(lines)


def build_mom_table_description(record: MeetingRecord) -> str:
    """Builds the MOM as an HTML table for Zoho Projects description field."""

    def h(s: str) -> str:
        """HTML-escape a string."""
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    # Header block
    meta_rows = [
        ("Meeting", record.meeting_title),
        ("Date", record.meeting_date),
        ("Place", record.place),
        ("Project", record.project_name),
    ]
    if record.mom_number:
        meta_rows.insert(0, ("MOM Ref", record.mom_number))
    if record.attendees:
        meta_rows.append(("Attended", ", ".join(record.attendees)))

    meta_html = "".join(
        f"<tr><td><b>{h(k)}</b></td><td>{h(v)}</td></tr>"
        for k, v in meta_rows
    )

    # Discussion table
    row_style = ' style="background:#f9f9f9"'
    rows_html = ""
    for i, dp in enumerate(record.discussion_points, 1):
        style = row_style if i % 2 == 0 else ""
        rows_html += (
            f"<tr{style}>"
            f"<td>{i}</td>"
            f"<td>{h(dp.point_of_discussion)}</td>"
            f"<td>{h(dp.discipline_of_work)}</td>"
            f"<td>{h(dp.conclusion_or_remark)}</td>"
            f"<td>{h(dp.responsible_party)}</td>"
            f"<td>{h(dp.target_date)}</td>"
            f"<td>{h(dp.status)}</td>"
            f"</tr>"
        )

    th = '<th style="background:#2c3e50;color:#fff;padding:6px 10px;text-align:left">'
    table_html = (
        f'<table border="1" cellpadding="6" cellspacing="0" '
        f'style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;font-size:13px">'
        f"<thead><tr>"
        f"{th}Sr. No.</th>{th}Point of Discussion</th>{th}Discipline</th>"
        f"{th}Conclusion / Remark</th>{th}Responsible Party</th>"
        f"{th}Target Date</th>{th}Status</th>"
        f"</tr></thead>"
        f"<tbody>{rows_html}</tbody>"
        f"</table>"
    )

    next_meeting = ""
    if record.next_meeting_date or record.next_meeting_place:
        next_meeting = (
            f"<p><b>Next Meeting:</b> {h(record.next_meeting_date)} "
            f"at {h(record.next_meeting_place)}</p>"
        )

    return (
        f'<table border="0" cellpadding="4" style="font-family:Arial,sans-serif;font-size:13px;margin-bottom:12px">'
        f"{meta_html}</table>"
        f"{table_html}"
        f"{next_meeting}"
    )


def generate_ai_key_points(record: MeetingRecord, api_key: str) -> str:
    """Uses OpenAI to produce a bulleted key-points summary of the meeting."""
    if not api_key:
        return _fallback_key_points(record)
    try:
        points_text = "\n".join(
            f"- [{dp.discipline_of_work}] {dp.point_of_discussion}"
            + (f" → {dp.conclusion_or_remark}" if dp.conclusion_or_remark else "")
            + (f" (Owner: {dp.responsible_party}, Due: {dp.target_date}, Status: {dp.status})" if dp.responsible_party else "")
            for dp in record.discussion_points
        )
        prompt = (
            f"You are a construction project manager's assistant. "
            f"From the meeting discussion points below, extract EXACTLY 5 to 7 of the most important ACTIONABLE decisions or tasks. "
            f"Output only a plain bulleted list using '• ' as the bullet. "
            f"Each bullet must be one concise sentence (max 15 words). No headers, no categories, no extra text. "
            f"Total response must be under 900 characters.\n\n"
            f"Meeting: {record.meeting_title} | Date: {record.meeting_date} | Project: {record.project_name}\n\n"
            f"Discussion Points:\n{points_text}"
        )
        client = OpenAI(api_key=api_key)
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=300,
            temperature=0.3,
        )
        return resp.choices[0].message.content.strip()[:1000]
    except Exception:
        return _fallback_key_points(record)


def _fallback_key_points(record: MeetingRecord) -> str:
    """Fallback bulleted key points without AI — capped at 1000 chars."""
    lines = [f"Meeting Summary — {record.meeting_title} ({record.meeting_date})", ""]
    for dp in record.discussion_points:
        bullet = f"• [{dp.discipline_of_work}] {dp.point_of_discussion}"
        if dp.conclusion_or_remark:
            bullet += f" → {dp.conclusion_or_remark}"
        if dp.responsible_party:
            bullet += f" | Owner: {dp.responsible_party} | Due: {dp.target_date} | {dp.status}"
        lines.append(bullet)
    if record.next_meeting_date or record.next_meeting_place:
        lines += ["", f"Next Meeting: {record.next_meeting_date} at {record.next_meeting_place}".strip()]
    return "\n".join(lines)[:1000]


def _to_zoho_date(date_str: str) -> str:
    """Convert any common date string to Zoho's required YYYY-MM-DD format."""
    from datetime import datetime
    for fmt in ("%d %B %Y", "%d %b %Y", "%B %d, %Y", "%b %d, %Y", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return date_str  # return as-is if unparseable


def push_mom_to_zoho(
    record: MeetingRecord,
    project_id: str,
    excel_bytes: bytes,
    excel_filename: str,
    api_key: str = "",
) -> tuple[bool, str]:
    """Creates a MOM record in the Zoho Projects custom MOMs module and attaches the Excel.
    Returns (success, message). Never raises — all exceptions are caught."""
    client_id = _get_zoho_secret("ZOHO_CLIENT_ID")
    client_secret = _get_zoho_secret("ZOHO_CLIENT_SECRET")
    refresh_token = _get_zoho_secret("ZOHO_REFRESH_TOKEN")
    portal_id = _get_zoho_secret("ZOHO_PORTAL_ID") or ZOHO_PORTAL_ID_DEFAULT
    if not (client_id and client_secret and refresh_token):
        return False, "Zoho credentials not configured."
    try:
        module_api_name = _get_moms_module_api_name(portal_id)
        access_token = _get_zoho_access_token(client_id, client_secret, refresh_token)
        record_name = f"{record.mom_number or record.meeting_title} \u2014 {record.meeting_date}"

        headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
        base = f"https://projectsapi.zoho.in/api/v3/portal/{portal_id}/module/{module_api_name}"

        # description = full MOM table; minutes_of_meeting = AI-generated bullet summary
        description = build_mom_table_description(record)
        key_points = generate_ai_key_points(record, api_key)

        # Build payload — include mandatory layout fields
        # minutes_of_meeting is capped at 1000 chars by Zoho's field limit
        payload: dict = {
            "name": record_name,
            "description": description,
            "minutes_of_meeting": key_points[:1000],
            "project": {"id": project_id},
            "date_of_decision": _to_zoho_date(record.meeting_date),
        }

        # Create the MOMs entity; project field links it to the right project
        create_resp = requests.post(
            f"{base}/entities",
            headers=headers,
            json=payload,
            timeout=15,
        )
        if not create_resp.ok:
            return False, (
                f"Failed to create Zoho record [{module_api_name}] ({create_resp.status_code}): "
                f"{create_resp.text[:400]}"
            )
        data = create_resp.json()
        # Response shape: {"id": ...} (flat) OR {"entity": {"id": ...}} OR {"entities": [{...}]}
        record_id = (
            data.get("id")
            or (data.get("entity") or {}).get("id")
            or (data.get("entities") or [{}])[0].get("id")
        )
        if not record_id:
            return True, (
                f"MOM record created in Zoho but could not extract its ID — attachment skipped. "
                f"Response: {str(data)[:200]}"
            )

        # Step 1: Get project folder ID via v2 folders endpoint (res_id is the field)
        folder_id = None
        folders_resp = requests.get(
            f"https://projectsapi.zoho.in/restapi/portal/{portal_id}"
            f"/projects/{project_id}/folders/",
            headers=headers,
            timeout=15,
        )
        if folders_resp.ok:
            for f in (folders_resp.json().get("folders") or []):
                fid = f.get("res_id") or f.get("id")
                if fid:
                    folder_id = fid
                    break

        # Step 2: Try v3 global upload first; fall back to v2 project documents upload
        attachment_id = None
        folders_debug = f"folders_status={folders_resp.status_code}, folders_body={folders_resp.text[:300]}"
        if not folder_id:
            # No folder found — try v3 global portal upload (needs upload rule configured)
            up = requests.post(
                f"https://projectsapi.zoho.in/api/v3/portal/{portal_id}/attachments",
                headers=headers,
                files={"file": (excel_filename, excel_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=30,
            )
            if up.ok:
                d = up.json()
                attachment_id = d.get("id") or (d.get("attachments") or [{}])[0].get("id")
            else:
                return True, (
                    f"MOM record created (ID: {record_id}). "
                    f"File upload skipped — no project folder found [{folders_debug}] "
                    f"and global upload failed ({up.status_code}): {up.text[:300]}"
                )
        else:
            # Upload via v2 project documents API (uploaddoc field + folder_id required)
            up = requests.post(
                f"https://projectsapi.zoho.in/restapi/portal/{portal_id}"
                f"/projects/{project_id}/documents/",
                headers=headers,
                files={"uploaddoc": (excel_filename, excel_bytes,
                                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"folder_id": folder_id},
                timeout=30,
            )
            if not up.ok:
                return True, (
                    f"MOM record created (ID: {record_id}), but file upload failed "
                    f"({up.status_code}): {up.text[:400]}"
                )
            d = up.json()
            docs = d if isinstance(d, list) else (d.get("documents") or [])
            attachment_id = (docs[0].get("id") or docs[0].get("res_id")) if docs else None

        if not attachment_id:
            return True, (
                f"MOM record created (ID: {record_id}), file uploaded but could not "
                f"read attachment ID. Response: {str(d)[:300]}"
            )

        # Step 3: Associate attachment with the MOMs entity
        assoc_resp = requests.post(
            f"https://projectsapi.zoho.in/api/v3/portal/{portal_id}"
            f"/attachments/{attachment_id}/associate",
            headers=headers,
            json={"entity_id": record_id, "entity_type": module_api_name},
            timeout=15,
        )
        if not assoc_resp.ok:
            return True, (
                f"MOM record created (ID: {record_id}), file uploaded (ID: {attachment_id}), "
                f"but association failed ({assoc_resp.status_code}): {assoc_resp.text[:300]}"
            )
        return True, f"MOM record created in Zoho (ID: {record_id}) with Excel attached successfully."
    except Exception as exc:
        return False, f"Zoho push failed: {exc}"


def run_app() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

    # ── Branded header ──
    st.markdown(
        '<div class="app-header">'
        '<div class="app-header-icon">'
        '<svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">'
        '<path d="M9 2H15L19 6V20C19 21.1 18.1 22 17 22H7C5.9 22 5 21.1 5 20V4C5 2.9 5.9 2 7 2H9Z" '
        'stroke="white" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>'
        '<path d="M9 13H15M9 17H13M14 2V6H19" '
        'stroke="white" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>'
        '</svg></div>'
        '<div class="app-header-text">'
        "<h1>SRMD MOM Generator</h1>"
        "<p>Paste site-visit notes, confirm project details, and generate a formatted Minutes of Meeting.</p>"
        "</div></div>",
        unsafe_allow_html=True,
    )

    with st.sidebar:
        st.markdown("#### AI Configuration")
        api_key = get_api_key()
        model = st.selectbox("Model", MODEL_OPTIONS, index=MODEL_OPTIONS.index(DEFAULT_MODEL))
        if api_key:
            st.markdown("API key configured")
        else:
            st.warning("No API key found — fallback parser active.")

    left_col, right_col = st.columns([1, 1.2], gap="large")

    with left_col:
        st.markdown('<div class="section-header">Meeting Details</div>', unsafe_allow_html=True)
        existing_mom_upload = st.file_uploader(
            "Upload existing MOM Excel (optional)",
            type=["xlsx"],
            help="Use an older or pre-made MOM workbook to prefill the form and add context for the new output.",
        )

        extracted_context = ExistingMomContext(attendees=[])
        if existing_mom_upload is not None:
            try:
                upload_signature = (
                    existing_mom_upload.name,
                    existing_mom_upload.size,
                )
                if st.session_state.get("existing_mom_signature") != upload_signature:
                    extracted_context = extract_existing_mom_context(existing_mom_upload)
                    st.session_state["existing_mom_signature"] = upload_signature
                    if extracted_context.project_name:
                        set_project_state(extracted_context.project_name)
                    if extracted_context.meeting_title:
                        st.session_state["meeting_title_input"] = extracted_context.meeting_title
                    if extracted_context.meeting_date:
                        st.session_state["meeting_date_input"] = extracted_context.meeting_date
                    if extracted_context.place:
                        st.session_state["place_input"] = extracted_context.place
                    if extracted_context.attendees:
                        st.session_state["attendees_input"] = "\n".join(extracted_context.attendees)
                    if extracted_context.discussion_notes and not st.session_state.get("meeting_notes_input", "").strip():
                        st.session_state["meeting_notes_input"] = extracted_context.discussion_notes
                    if extracted_context.workbook_text and not st.session_state.get("extra_context_input", "").strip():
                        st.session_state["extra_context_input"] = (
                            "Existing MOM workbook context:\n" + extracted_context.workbook_text
                        )
                else:
                    extracted_context = extract_existing_mom_context(existing_mom_upload)
                st.caption("Existing MOM loaded. Any extracted values are prefilled below and can still be edited.")
            except Exception as exc:
                st.warning(f"Could not read the uploaded MOM Excel file: {exc}")

        st.session_state.setdefault("project_name_select", "Select project")
        st.session_state.setdefault("project_name_ngh_multi", [])
        st.session_state.setdefault("project_name_custom_input", "")
        st.session_state.setdefault("meeting_title_input", "Site Visit Meeting")

        _zoho_opts, _zoho_err = fetch_zoho_project_options()
        if _zoho_err:
            st.warning(f"Could not load Zoho projects: {_zoho_err}")
        _project_selection_options = ["Select project"] + _zoho_opts + [MANUAL_PROJECT_OPTION]
        selected_project_name = st.selectbox(
            "Project name",
            _project_selection_options,
            key="project_name_select",
        )
        project_name = selected_project_name
        if selected_project_name in NGH_PROJECT_OPTIONS:
            default_ngh_projects = st.session_state.get("project_name_ngh_multi") or [selected_project_name]
            selected_ngh_projects = st.multiselect(
                "Select NGH projects",
                NGH_PROJECT_OPTIONS,
                default=default_ngh_projects,
                key="project_name_ngh_multi",
            )
            project_name = ", ".join(selected_ngh_projects)
        elif selected_project_name == MANUAL_PROJECT_OPTION:
            project_name = st.text_input(
                "Manual project name",
                key="project_name_custom_input",
                placeholder="Enter project name",
            )
        elif selected_project_name == "Select project":
            project_name = ""
        meeting_title = st.text_input("Meeting title", key="meeting_title_input", value="Site Visit Meeting")
        mom_number = st.text_input(
            "MOM Reference No.",
            key="mom_number_input",
            placeholder="e.g. MOM-003",
        )
        meeting_date = st.text_input(
            "Meeting date",
            key="meeting_date_input",
            placeholder="e.g. 14 March 2026",
        )
        place = st.text_input("Place", key="place_input", placeholder="e.g. Ahmedabad site office")
        attendees_text = st.text_area(
            "Attendees",
            key="attendees_input",
            placeholder="One attendee per line\nJohn Shah - Consultant\nMehul Patel - Contractor",
            height=140,
        )
        extra_context = st.text_area(
            "Optional context",
            key="extra_context_input",
            placeholder="Anything else the AI should know, like phase, package, contractor names, or purpose of the visit.",
            height=120,
        )
        with st.expander("Next meeting details (optional)"):
            next_meeting_date = st.text_input(
                "Next meeting date",
                key="next_meeting_date_input",
                placeholder="e.g. 28 March 2026",
            )
            next_meeting_place = st.text_input(
                "Next meeting place",
                key="next_meeting_place_input",
                placeholder="e.g. Ahmedabad site office",
            )

    with right_col:
        st.markdown('<div class="section-header">Meeting Notes</div>', unsafe_allow_html=True)
        meeting_notes = st.text_area(
            "Paste the raw minutes / site visit notes here",
            key="meeting_notes_input",
            placeholder="Paste the long MOM text, WhatsApp notes, bullet points, or site visit summary here.",
            height=420,
        )

    st.markdown("")  # spacing
    template_ready = os.path.exists(DEFAULT_TEMPLATE_PATH)
    generate_disabled = not meeting_notes.strip() or not template_ready
    gen_col1, gen_col2, gen_col3 = st.columns([1, 1, 1])
    with gen_col2:
        generate_clicked = st.button(
            "Generate MOM", type="primary", disabled=generate_disabled, use_container_width=True
        )

    if not template_ready:
        st.warning(
            f"Template not found at `{DEFAULT_TEMPLATE_PATH}`. Add the SRMD MOM template to the repo root to continue."
        )

    if generate_clicked:
        try:
            with st.spinner("Structuring the meeting notes and preparing the Excel file..."):
                uploaded_mom_notes = ""
                uploaded_mom_context = ""
                uploaded_mom_attendees: list[str] = []
                if existing_mom_upload is not None:
                    existing_context = extract_existing_mom_context(existing_mom_upload)
                    uploaded_mom_notes = existing_context.discussion_notes
                    uploaded_mom_context = existing_context.workbook_text
                    uploaded_mom_attendees = existing_context.attendees or []

                combined_notes = meeting_notes.strip()
                if uploaded_mom_notes:
                    combined_notes = f"{combined_notes}\n\nReference MOM Excel notes:\n{uploaded_mom_notes}".strip()

                combined_context = extra_context.strip()
                if uploaded_mom_context:
                    addition = f"Reference MOM Excel workbook details:\n{uploaded_mom_context}"
                    combined_context = f"{combined_context}\n\n{addition}".strip() if combined_context else addition

                combined_attendees = unique_nonempty(
                    clean_lines(attendees_text)
                    + uploaded_mom_attendees
                    + extract_attendees_from_text(attendees_text, meeting_notes, combined_context)
                )
                attendees_payload = "\n".join(combined_attendees)

                record = generate_meeting_record(
                    api_key=api_key.strip(),
                    model=model,
                    project_name=project_name,
                    meeting_title=meeting_title,
                    meeting_date=meeting_date,
                    place=place,
                    attendees_text=attendees_payload,
                    notes=combined_notes,
                    extra_context=combined_context,
                    mom_number=mom_number.strip(),
                    next_meeting_date=next_meeting_date.strip(),
                    next_meeting_place=next_meeting_place.strip(),
                )
                record.attendees = unique_nonempty(record.attendees + combined_attendees)
                generated_workbook = build_workbook(template_bytes(), record)
                generated_pdf = build_pdf_report(record)

            st.success("Excel MOM generated successfully.")

            # Persist generated data so preview/download/Zoho button survive re-renders
            _excel_fname = sanitize_filename(
                f"{record.project_name}_{record.meeting_date}_MOM"
                + (f"_{sanitize_filename(record.mom_number)}" if record.mom_number else "")
            ) + ".xlsx"
            st.session_state["_mom_record"] = record
            st.session_state["_mom_workbook"] = generated_workbook
            st.session_state["_mom_pdf"] = generated_pdf
            st.session_state["_mom_excel_fname"] = _excel_fname
            st.session_state["_mom_project_name"] = project_name
            st.session_state["_zoho_push_result"] = None  # reset on new generation

        except Exception as exc:
            st.error(f"Could not generate the MOM workbook: {exc}")

    # --- Persistent results section (survives tab clicks and re-renders) ---
    _r: MeetingRecord | None = st.session_state.get("_mom_record")
    if _r is not None:
        _wb: bytes = st.session_state["_mom_workbook"]
        _pdf: bytes = st.session_state["_mom_pdf"]
        _excel_fname: str = st.session_state["_mom_excel_fname"]
        _proj_name: str = st.session_state.get("_mom_project_name", "")

        _STATUS_STYLE_MAP = {
            "Open": "background-color: #ffe0e0",
            "In Progress": "background-color: #fff3cd",
            "Closed": "background-color: #d4edda",
            "Deferred": "background-color: #e2e3e5",
        }

        preview_rows = [
            {
                "Sr.": index,
                "Point of Discussion": point.point_of_discussion,
                "Discipline": point.discipline_of_work,
                "Status": point.status,
                "Owner": point.responsible_party,
                "Due Date": point.target_date,
                "Conclusion / Remark": point.conclusion_or_remark,
            }
            for index, point in enumerate(_r.discussion_points, start=1)
        ]

        excel_tab, pdf_tab, email_tab = st.tabs(["Excel Preview", "PDF Preview", "Email Draft"])

        with excel_tab:
            st.markdown('<div class="section-header">Structured Preview</div>', unsafe_allow_html=True)
            meta_col1, meta_col2 = st.columns(2)
            with meta_col1:
                st.markdown(f"**Project:** {_r.project_name}")
                st.markdown(f"**Meeting:** {_r.meeting_title}")
                if _r.mom_number:
                    st.markdown(f"**MOM Ref:** {_r.mom_number}")
            with meta_col2:
                st.markdown(f"**Date:** {_r.meeting_date}")
                st.markdown(f"**Place:** {_r.place}")
                if _r.next_meeting_date or _r.next_meeting_place:
                    next_info = " — ".join(filter(None, [_r.next_meeting_date, _r.next_meeting_place]))
                    st.markdown(f"**Next Meeting:** {next_info}")
            with st.expander(f"Attendees ({len(_r.attendees)})"):
                st.markdown(", ".join(_r.attendees))
            df = pd.DataFrame(preview_rows)
            try:
                styled_df = df.style.map(
                    lambda v: _STATUS_STYLE_MAP.get(v, ""), subset=["Status"]
                )
                st.dataframe(styled_df, use_container_width=True, hide_index=True)
            except Exception:
                st.dataframe(df, use_container_width=True, hide_index=True)

        with pdf_tab:
            st.markdown('<div class="section-header">PDF Preview</div>', unsafe_allow_html=True)
            pdf_base64 = base64.b64encode(_pdf).decode("ascii")
            st.markdown(
                (
                    "<iframe style='width:100%; height:900px; border:1px solid #ddd;' "
                    f"src='data:application/pdf;base64,{pdf_base64}'></iframe>"
                ),
                unsafe_allow_html=True,
            )

        with email_tab:
            st.markdown('<div class="section-header">Email Draft</div>', unsafe_allow_html=True)
            st.caption("Copy and paste this into your email client. Edit as needed before sending.")
            st.code(build_email_draft(_r), language="")

        mom_suffix = f"_{sanitize_filename(_r.mom_number)}" if _r.mom_number else ""
        output_name = sanitize_filename(f"{_r.project_name}_{_r.meeting_date}_MOM{mom_suffix}") + ".xlsx"
        pdf_name = sanitize_filename(f"{_r.project_name}_{_r.meeting_date}_MOM{mom_suffix}") + ".pdf"

        st.divider()
        st.markdown('<div class="section-header">Actions</div>', unsafe_allow_html=True)
        dl_col1, dl_col2, zoho_col = st.columns([1, 1, 1])
        with dl_col1:
            st.download_button(
                label="Download Excel",
                data=_wb,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with dl_col2:
            st.download_button(
                label="Download PDF",
                data=_pdf,
                file_name=pdf_name,
                mime="application/pdf",
                use_container_width=True,
            )

        with zoho_col:
            _zoho_pid = get_zoho_project_id(_proj_name) if _proj_name else None
            _prev_result: tuple[bool, str] | None = st.session_state.get("_zoho_push_result")
            if _prev_result is not None:
                if _prev_result[0]:
                    st.success(f"Zoho: {_prev_result[1]}")
                else:
                    st.warning(f"Zoho push: {_prev_result[1]}")

            if _zoho_pid:
                if st.button("Push to Zoho Projects", type="primary", use_container_width=True):
                    with st.spinner("Pushing MOM to Zoho Projects..."):
                        _ok, _msg = push_mom_to_zoho(
                            record=_r,
                            project_id=_zoho_pid,
                            excel_bytes=_wb,
                            excel_filename=_excel_fname,
                            api_key=get_api_key(),
                        )
                    st.session_state["_zoho_push_result"] = (_ok, _msg)
                    st.rerun()
            else:
                st.info(
                    "Zoho push unavailable for this project.\n\n"
                    "Select a **Projects - \\*** project from the dropdown to enable it."
                )


if __name__ == "__main__":
    run_app()
