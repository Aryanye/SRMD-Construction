from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from io import BytesIO
import json
import os
from pathlib import Path
import re
from typing import Any

from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
import streamlit as st

APP_TITLE = "SRMD MOM Generator"
DEFAULT_MODEL = "gpt-5.1"
MODEL_OPTIONS = ["gpt-5.1", "gpt-5-mini", "gpt-4.1"]
BASE_DIR = Path(__file__).resolve().parent
HEAD
DEFAULT_TEMPLATE_PATH = "SRMD MOM Format.xlsx"
DEFAULT_TEMPLATE_PATH = BASE_DIR / "SRMD MOM Format.xlsx"
4b2ed8a (Update MOM template path)
DISCUSSION_START_ROW = 18
BASE_DISCUSSION_ROWS = 9
FOOTER_START_ROW = 27
MIDDLE_DISCUSSION_TEMPLATE_ROW = 25

SYSTEM_PROMPT = """
You convert raw site-visit meeting notes into a structured Minutes of Meeting record.

Rules:
- Prefer any user-supplied field over inference.
- Infer missing details only from the notes and optional context.
- Keep wording professional, concise, and suitable for an Excel MOM register.
- Split the discussion into clear action-oriented points.
- Return strict JSON only.
- If a field is genuinely unclear, use a sensible placeholder instead of leaving it blank.
""".strip()


@dataclass
class DiscussionPoint:
    point_of_discussion: str
    discipline_of_work: str
    conclusion_or_remark: str


@dataclass
class MeetingRecord:
    project_name: str
    meeting_title: str
    meeting_date: str
    place: str
    attendees: list[str]
    discussion_points: list[DiscussionPoint]


def sanitize_text(value: Any, fallback: str) -> str:
    text = str(value or "").strip()
    return text or fallback


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return cleaned.strip("._") or "mom_report"


def get_api_key() -> str:
    secret_value = st.secrets.get("OPENAI_API_KEY", "")
    if secret_value:
        return str(secret_value)
    return os.getenv("OPENAI_API_KEY", "")


def clean_lines(value: str) -> list[str]:
    return [line.strip(" -\t") for line in value.splitlines() if line.strip()]


def parse_json_response(raw_text: str) -> dict[str, Any]:
    raw_text = raw_text.strip()
    if not raw_text:
        raise ValueError("The model returned an empty response.")

    try:
        return json.loads(raw_text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", raw_text, re.DOTALL)
        if not match:
            raise ValueError("The model response was not valid JSON.")
        return json.loads(match.group(0))


def infer_discipline(text: str) -> str:
    lowered = text.lower()
    mapping = {
        "Civil": ["concrete", "slab", "column", "beam", "plaster", "brick", "masonry", "floor", "excavation"],
        "MEP": ["electrical", "wiring", "cable", "lighting", "plumbing", "drain", "pipe", "hvac", "fire fighting"],
        "Architecture": ["finish", "façade", "facade", "paint", "tile", "door", "window", "ceiling", "elevation"],
        "Safety": ["safety", "ppe", "barricade", "housekeeping", "hazard", "incident"],
        "Planning": ["schedule", "timeline", "delay", "handover", "approval", "submission", "procurement"],
    }
    for discipline, keywords in mapping.items():
        if any(keyword in lowered for keyword in keywords):
            return discipline
    return "General"


def heuristic_discussion_points(notes: str) -> list[DiscussionPoint]:
    lines = clean_lines(notes)
    if not lines:
        lines = [segment.strip() for segment in re.split(r"(?<=[.!?])\s+", notes) if segment.strip()]

    points: list[DiscussionPoint] = []
    for line in lines:
        normalized = re.sub(r"^\d+[\).:-]?\s*", "", line).strip()
        if not normalized:
            continue
        conclusion = "Review on site and close the action as discussed."
        if any(token in normalized.lower() for token in ["approved", "completed", "closed"]):
            conclusion = "Noted as completed / accepted during the meeting."
        elif any(token in normalized.lower() for token in ["provide", "submit", "share", "revise", "rectify"]):
            conclusion = "Concerned team to take action and update in the next review."
        points.append(
            DiscussionPoint(
                point_of_discussion=normalized,
                discipline_of_work=infer_discipline(normalized),
                conclusion_or_remark=conclusion,
            )
        )

    return points[:25] or [
        DiscussionPoint(
            point_of_discussion="Review the submitted site visit notes and confirm the action items.",
            discipline_of_work="General",
            conclusion_or_remark="Action owners to align and issue an updated status after the meeting.",
        )
    ]


def heuristic_meeting_record(
    project_name: str,
    meeting_title: str,
    meeting_date: str,
    place: str,
    attendees_text: str,
    notes: str,
) -> MeetingRecord:
    attendees = clean_lines(attendees_text)
    return MeetingRecord(
        project_name=sanitize_text(project_name, "PROJECT NAME TO BE CONFIRMED"),
        meeting_title=sanitize_text(meeting_title, "Site Visit Meeting"),
        meeting_date=sanitize_text(meeting_date, "Date to be confirmed"),
        place=sanitize_text(place, "Place to be confirmed"),
        attendees=attendees or ["Attendee details to be confirmed from the meeting notes."],
        discussion_points=heuristic_discussion_points(notes),
    )


def build_prompt(
    project_name: str,
    meeting_title: str,
    meeting_date: str,
    place: str,
    attendees_text: str,
    notes: str,
    extra_context: str,
) -> str:
    payload = {
        "user_inputs": {
            "project_name": project_name.strip(),
            "meeting_title": meeting_title.strip(),
            "meeting_date": meeting_date.strip(),
            "place": place.strip(),
            "attendees": clean_lines(attendees_text),
        },
        "extra_context": extra_context.strip(),
        "meeting_notes": notes.strip(),
        "output_schema": {
            "project_name": "string",
            "meeting_title": "string",
            "meeting_date": "string",
            "place": "string",
            "attendees": ["short attendee strings"],
            "discussion_points": [
                {
                    "point_of_discussion": "string under 180 characters",
                    "discipline_of_work": "short label like Civil / MEP / Architecture / Safety / Planning / General",
                    "conclusion_or_remark": "string under 160 characters",
                }
            ],
        },
    }
    return json.dumps(payload, indent=2)


def normalize_meeting_record(payload: dict[str, Any]) -> MeetingRecord:
    raw_points = payload.get("discussion_points")
    points: list[DiscussionPoint] = []
    if isinstance(raw_points, list):
        for item in raw_points:
            if not isinstance(item, dict):
                continue
            points.append(
                DiscussionPoint(
                    point_of_discussion=sanitize_text(
                        item.get("point_of_discussion"),
                        "Discussion point to be confirmed.",
                    ),
                    discipline_of_work=sanitize_text(item.get("discipline_of_work"), "General"),
                    conclusion_or_remark=sanitize_text(
                        item.get("conclusion_or_remark"),
                        "Action owner to review and update.",
                    ),
                )
            )

    attendees_value = payload.get("attendees")
    attendees = []
    if isinstance(attendees_value, list):
        attendees = [str(item).strip() for item in attendees_value if str(item).strip()]

    return MeetingRecord(
        project_name=sanitize_text(payload.get("project_name"), "PROJECT NAME TO BE CONFIRMED"),
        meeting_title=sanitize_text(payload.get("meeting_title"), "Site Visit Meeting"),
        meeting_date=sanitize_text(payload.get("meeting_date"), "Date to be confirmed"),
        place=sanitize_text(payload.get("place"), "Place to be confirmed"),
        attendees=attendees or ["Attendee details to be confirmed from the meeting notes."],
        discussion_points=points
        or [
            DiscussionPoint(
                point_of_discussion="Review the submitted site visit notes and confirm the action items.",
                discipline_of_work="General",
                conclusion_or_remark="Action owners to align and issue an updated status after the meeting.",
            )
        ],
    )


def generate_meeting_record(
    api_key: str,
    model: str,
    project_name: str,
    meeting_title: str,
    meeting_date: str,
    place: str,
    attendees_text: str,
    notes: str,
    extra_context: str,
) -> MeetingRecord:
    fallback = heuristic_meeting_record(
        project_name=project_name,
        meeting_title=meeting_title,
        meeting_date=meeting_date,
        place=place,
        attendees_text=attendees_text,
        notes=notes,
    )
    if not api_key:
        return fallback

    client = OpenAI(api_key=api_key)
    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "system",
                "content": [{"type": "input_text", "text": SYSTEM_PROMPT}],
            },
            {
                "role": "user",
                "content": [{"type": "input_text", "text": build_prompt(project_name, meeting_title, meeting_date, place, attendees_text, notes, extra_context)}],
            },
        ],
    )
    return normalize_meeting_record(parse_json_response(response.output_text))


def template_bytes_from_upload(uploaded_template: Any) -> bytes:
    if uploaded_template is not None:
        return uploaded_template.getvalue()
    with open(DEFAULT_TEMPLATE_PATH, "rb") as template_file:
        return template_file.read()


def copy_row_style(ws: Any, source_row: int, target_row: int, start_col: int = 1, end_col: int = 10) -> None:
    source_height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].height = source_height
    for col in range(start_col, end_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if isinstance(source, MergedCell):
            continue
        target._style = copy(source._style)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)


def rebuild_dynamic_layout(ws: Any, extra_rows: int) -> None:
    if extra_rows <= 0:
        return

    saved_heights = {row: ws.row_dimensions[row].height for row in range(FOOTER_START_ROW, 34)}

    ws.move_range(f"B{FOOTER_START_ROW}:F33", rows=extra_rows, cols=0, translate=True)

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= FOOTER_START_ROW:
            ws.merged_cells.ranges.remove(merged_range)

    for row in range(FOOTER_START_ROW, FOOTER_START_ROW + extra_rows):
        copy_row_style(ws, MIDDLE_DISCUSSION_TEMPLATE_ROW, row, start_col=2, end_col=6)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

    for row in range(FOOTER_START_ROW + extra_rows, 34 + extra_rows):
        original_row = row - extra_rows
        if original_row in saved_heights:
            ws.row_dimensions[row].height = saved_heights[original_row]

    ws.merge_cells(start_row=28 + extra_rows, start_column=2, end_row=28 + extra_rows, end_column=6)
    ws.merge_cells(start_row=29 + extra_rows, start_column=2, end_row=29 + extra_rows, end_column=6)
    ws.merge_cells(start_row=30 + extra_rows, start_column=2, end_row=32 + extra_rows, end_column=6)
    ws.merge_cells(start_row=33 + extra_rows, start_column=3, end_row=33 + extra_rows, end_column=4)


def fill_attendees(ws: Any, attendees: list[str]) -> None:
    attendee_cells = ["C8", "D9", "D10", "D11", "D12", "D13", "D14"]
    for coord, attendee in zip(attendee_cells, attendees[: len(attendee_cells)]):
        ws[coord] = attendee
        ws[coord].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if len(attendees) > len(attendee_cells):
        overflow = "Additional attendees: " + ", ".join(attendees[len(attendee_cells) :])
        ws["D14"] = f"{ws['D14'].value}\n{overflow}" if ws["D14"].value else overflow
        ws["D14"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def fill_discussion_table(ws: Any, discussion_points: list[DiscussionPoint]) -> None:
    points = discussion_points[:25]
    extra_rows = max(0, len(points) - BASE_DISCUSSION_ROWS)
    rebuild_dynamic_layout(ws, extra_rows)

    for index, point in enumerate(points, start=DISCUSSION_START_ROW):
        serial_no = index - DISCUSSION_START_ROW + 1
        ws[f"B{index}"] = serial_no
        ws[f"C{index}"] = point.point_of_discussion
        ws[f"E{index}"] = point.discipline_of_work
        ws[f"F{index}"] = point.conclusion_or_remark

        for coord in [f"C{index}", f"E{index}", f"F{index}"]:
            ws[coord].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_workbook(template_content: bytes, meeting_record: MeetingRecord) -> bytes:
    workbook = load_workbook(BytesIO(template_content))
    worksheet = workbook[workbook.sheetnames[0]]

    worksheet["B2"] = meeting_record.project_name.upper()
    worksheet["D4"] = meeting_record.meeting_title
    worksheet["D5"] = f"{meeting_record.meeting_date}\nPlace: {meeting_record.place}"
    worksheet["D5"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    fill_attendees(worksheet, meeting_record.attendees)
    fill_discussion_table(worksheet, meeting_record.discussion_points)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

def run_app() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    st.title(APP_TITLE)
    st.caption(
        "Paste your site-visit meeting notes, confirm the basic project details, and download an Excel MOM "
        "that follows your SRMD template format."
    )

    with st.sidebar:
        st.header("AI Settings")
        api_key = st.text_input(
            "OpenAI API key",
            value=get_api_key(),
            type="password",
            help="Loads from Streamlit Secrets when deployed, or from your local environment when running locally.",
        )
        model = st.selectbox("Model", MODEL_OPTIONS, index=MODEL_OPTIONS.index(DEFAULT_MODEL))
        st.info(
            "If the API key is blank, the app still works with a basic local parser, but AI will do a better job "
            "inferring missing fields and cleaning up discussion points."
        )

    left_col, right_col = st.columns([1, 1.2])

    with left_col:
        st.subheader("Meeting Inputs")
        project_name = st.text_input("Project name", placeholder="e.g. SRMD Warehouse Extension")
        meeting_title = st.text_input("Meeting title", value="Site Visit Meeting")
        meeting_date = st.text_input("Meeting date", placeholder="e.g. 14 March 2026")
        place = st.text_input("Place", placeholder="e.g. Ahmedabad site office")
        attendees_text = st.text_area(
            "Attendees",
            placeholder="One attendee per line\nJohn Shah - Consultant\nMehul Patel - Contractor",
            height=140,
        )
        extra_context = st.text_area(
            "Optional context",
            placeholder="Anything else the AI should know, like phase, package, contractor names, or purpose of the visit.",
            height=120,
        )
        uploaded_template = st.file_uploader(
            "Excel template (optional)",
            type=["xlsx"],
            help="Leave this empty to use the SRMD MOM template stored in the repo.",
        )

    with right_col:
        st.subheader("Meeting Notes")
        meeting_notes = st.text_area(
            "Paste the raw minutes / site visit notes here",
            placeholder="Paste the long MOM text, WhatsApp notes, bullet points, or site visit summary here.",
            height=420,
        )

    template_ready = uploaded_template is not None or os.path.exists(DEFAULT_TEMPLATE_PATH)
    generate_disabled = not meeting_notes.strip() or not template_ready
    generate_clicked = st.button("Generate Excel MOM", type="primary", disabled=generate_disabled)

    if not template_ready:
        st.warning(
            f"Template not found at `{DEFAULT_TEMPLATE_PATH}`. Upload the template file above to continue."
        )

    if generate_clicked:
        try:
            with st.spinner("Structuring the meeting notes and preparing the Excel file..."):
                record = generate_meeting_record(
                    api_key=api_key.strip(),
                    model=model,
                    project_name=project_name,
                    meeting_title=meeting_title,
                    meeting_date=meeting_date,
                    place=place,
                    attendees_text=attendees_text,
                    notes=meeting_notes,
                    extra_context=extra_context,
                )
                generated_workbook = build_workbook(template_bytes_from_upload(uploaded_template), record)

            st.success("Excel MOM generated successfully.")

            preview_rows = [
                {
                    "Sr. No.": index,
                    "Point of discussion": point.point_of_discussion,
                    "Discipline": point.discipline_of_work,
                    "Conclusion / Remark": point.conclusion_or_remark,
                }
                for index, point in enumerate(record.discussion_points, start=1)
            ]

            st.subheader("Structured Preview")
            st.write(f"**Project:** {record.project_name}")
            st.write(f"**Meeting:** {record.meeting_title}")
            st.write(f"**Date:** {record.meeting_date}")
            st.write(f"**Place:** {record.place}")
            st.write("**Attendees:**")
            for attendee in record.attendees:
                st.write(f"- {attendee}")
            st.dataframe(preview_rows, use_container_width=True, hide_index=True)

            output_name = sanitize_filename(f"{record.project_name}_{record.meeting_date}_MOM") + ".xlsx"
            st.download_button(
                label="Download Excel File",
                data=generated_workbook,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            st.error(f"Could not generate the MOM workbook: {exc}")


if __name__ == "__main__":
    run_app()
