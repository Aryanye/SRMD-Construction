from __future__ import annotations

from datetime import datetime
from pathlib import Path
import math

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

SOURCE_XLSX = Path("V4 ANA Feasibility Financial Model 11.06.25 CS.xlsx")
REFINED_XLSX = Path("Refined_Financial_Model_Working.xlsx")
SUMMARY_PDF = Path("Investment_Summary_Report.pdf")


def safe_num(v, default=0.0):
    if v is None:
        return default
    try:
        return float(v)
    except Exception:
        return default


def irr(cashflows):
    # Cashflows assumed periodic, t=0..n-1. Returns None if no sign change.
    if not (any(cf > 0 for cf in cashflows) and any(cf < 0 for cf in cashflows)):
        return None

    def npv(rate):
        return sum(cf / ((1 + rate) ** t) for t, cf in enumerate(cashflows))

    lo, hi = -0.95, 5.0
    f_lo, f_hi = npv(lo), npv(hi)
    if f_lo * f_hi > 0:
        return None

    for _ in range(200):
        mid = (lo + hi) / 2
        f_mid = npv(mid)
        if abs(f_mid) < 1e-9:
            return mid
        if f_lo * f_mid <= 0:
            hi, f_hi = mid, f_mid
        else:
            lo, f_lo = mid, f_mid
    return mid


def build_model(params):
    years = [1, 2, 3, 4]
    sales_mix = params["sales_mix"]
    cons_mix = params["cons_mix"]
    approval_mix = params["approval_mix"]

    rows = []
    opening = 0.0
    for i, y in enumerate(years):
        resi_price = params["price_resi"] * ((1 + params["sales_growth"]) ** i) * params["revenue_multiplier"]
        comm_price = params["price_comm"] * ((1 + params["sales_growth"]) ** i) * params["revenue_multiplier"]
        gross_rev = (
            params["resi_area"] * sales_mix[i] * resi_price + params["comm_area"] * sales_mix[i] * comm_price
        ) / 1e7

        lo_payout = gross_rev * params["lo_share"]
        net_rev = gross_rev - lo_payout
        cons_out = params["construction"] * params["cost_multiplier"] * cons_mix[i]
        appr_out = params["approval"] * params["cost_multiplier"] * approval_mix[i]
        smoh_out = gross_rev * params["smoh_rate"]
        stamp = params["stamp_duty"] if i == 0 else 0.0
        dep_out = params["deposit"] if i == 0 else 0.0
        dep_ref = params["deposit"] if i == 3 else 0.0

        pre_int = net_rev - cons_out - appr_out - smoh_out - stamp - dep_out + dep_ref
        interest = max(-opening, 0.0) * params["interest_rate"]
        net_cf = pre_int - interest
        closing = opening + net_cf

        rows.append(
            {
                "year": f"Y{y}",
                "sales_mix": sales_mix[i],
                "resi_price": resi_price,
                "comm_price": comm_price,
                "gross_rev": gross_rev,
                "lo_payout": lo_payout,
                "net_rev": net_rev,
                "cons_out": cons_out,
                "appr_out": appr_out,
                "smoh_out": smoh_out,
                "stamp": stamp,
                "dep_out": dep_out,
                "dep_ref": dep_ref,
                "pre_interest_cf": pre_int,
                "opening": opening,
                "interest": interest,
                "net_cf": net_cf,
                "closing": closing,
            }
        )
        opening = closing

    gross_total = sum(r["gross_rev"] for r in rows)
    net_total = sum(r["net_cf"] for r in rows)
    npv_15 = sum(r["net_cf"] / ((1 + 0.15) ** i) for i, r in enumerate(rows))
    model_irr = irr([r["net_cf"] for r in rows])

    negatives = sum(-r["net_cf"] for r in rows if r["net_cf"] < 0)
    positives = sum(r["net_cf"] for r in rows if r["net_cf"] > 0)
    equity_multiple = (positives / negatives) if negatives > 0 else None

    return {
        "rows": rows,
        "kpi": {
            "gross_revenue": gross_total,
            "net_cashflow": net_total,
            "npv_15": npv_15,
            "irr": model_irr,
            "equity_multiple": equity_multiple,
            "total_interest": sum(r["interest"] for r in rows),
            "total_cost": sum(r["cons_out"] + r["appr_out"] + r["smoh_out"] + r["stamp"] + r["interest"] for r in rows),
        },
    }


def fmt_cr(x):
    return f"{x:,.2f} Cr"


def fmt_pct(x):
    if x is None:
        return "NA"
    return f"{x*100:.2f}%"


def main():
    wb_values = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)

    params = {
        "land_area_sqm": safe_num(wb_values["Input"]["C4"].value),
        "project_start": wb_values["Input"]["C5"].value,
        "project_period_years": int(safe_num(wb_values["Input"]["C6"].value, 4.0)),
        "price_resi": safe_num(wb_values["Input"]["C7"].value),
        "price_comm": safe_num(wb_values["Input"]["C8"].value),
        "sales_growth": safe_num(wb_values["Input"]["C11"].value),
        "resi_area": safe_num(wb_values["FSI & Area Cal"]["E48"].value),
        "comm_area": safe_num(wb_values["FSI & Area Cal"]["E49"].value),
        "lo_share": safe_num(wb_values["ANA Summary"]["E46"].value),
        "deposit": safe_num(wb_values["ANA Summary"]["F48"].value),
        "approval": safe_num(wb_values["Approvals & Liasoning"]["G22"].value) / 1e7,
        "construction": safe_num(wb_values["Detailed Costing"]["E32"].value),
        "stamp_duty": safe_num(wb_values["Input"]["B17"].value),
        "smoh_rate": safe_num(wb_values["ANA Summary"]["E34"].value, 0.07),
        "interest_rate": 0.12,
        "sales_mix": [
            safe_num(wb_values["Sales Phasing"]["D2"].value),
            safe_num(wb_values["Sales Phasing"]["D6"].value),
            safe_num(wb_values["Sales Phasing"]["D10"].value),
            safe_num(wb_values["Sales Phasing"]["D14"].value),
        ],
        "cons_mix": [
            safe_num(wb_values["Cash Flow"]["E6"].value),
            safe_num(wb_values["Cash Flow"]["F6"].value),
            safe_num(wb_values["Cash Flow"]["G6"].value),
            safe_num(wb_values["Cash Flow"]["H6"].value),
        ],
        "approval_mix": [0.45, 0.25, 0.20, 0.10],
        "cost_multiplier": 1.0,
        "revenue_multiplier": 1.0,
    }

    # Normalize phasing if source is imperfect.
    sales_total = sum(params["sales_mix"])
    if sales_total <= 0:
        params["sales_mix"] = [0.25, 0.25, 0.25, 0.25]
    elif abs(sales_total - 1.0) > 1e-8:
        params["sales_mix"] = [x / sales_total for x in params["sales_mix"]]

    cons_total = sum(params["cons_mix"])
    if cons_total <= 0:
        params["cons_mix"] = [0.25, 0.25, 0.25, 0.25]
    elif abs(cons_total - 1.0) > 1e-8:
        params["cons_mix"] = [x / cons_total for x in params["cons_mix"]]

    base = build_model(params)

    scenarios = {
        "Base": {"sales_growth": params["sales_growth"], "revenue_multiplier": 1.00, "cost_multiplier": 1.00},
        "Upside": {"sales_growth": params["sales_growth"] + 0.02, "revenue_multiplier": 1.03, "cost_multiplier": 0.95},
        "Downside": {"sales_growth": max(params["sales_growth"] - 0.02, 0.0), "revenue_multiplier": 0.95, "cost_multiplier": 1.10},
    }

    scenario_rows = []
    for name, adj in scenarios.items():
        p = dict(params)
        p.update(adj)
        out = build_model(p)
        scenario_rows.append(
            [
                name,
                fmt_cr(out["kpi"]["gross_revenue"]),
                fmt_cr(out["kpi"]["net_cashflow"]),
                fmt_cr(out["kpi"]["npv_15"]),
                fmt_pct(out["kpi"]["irr"]),
            ]
        )

    # Build refined workbook
    wb_new = Workbook()
    ws_a = wb_new.active
    ws_a.title = "Assumptions"
    ws_m = wb_new.create_sheet("Annual_Model")
    ws_s = wb_new.create_sheet("Scenario_Analysis")
    ws_d = wb_new.create_sheet("Model_Diagnostics")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    assumptions = [
        ("Project Start", params["project_start"]),
        ("Project Period (Years)", params["project_period_years"]),
        ("Land Area (SqM)", params["land_area_sqm"]),
        ("Residential Area (Sft)", params["resi_area"]),
        ("Commercial Area (Sft)", params["comm_area"]),
        ("Start Price - Residential (Rs/sft)", params["price_resi"]),
        ("Start Price - Commercial (Rs/sft)", params["price_comm"]),
        ("Sales Growth YoY", params["sales_growth"]),
        ("Landowner Revenue Share", params["lo_share"]),
        ("Refundable Deposit (Cr)", params["deposit"]),
        ("Approval Costs (Cr)", params["approval"]),
        ("Construction Costs (Cr)", params["construction"]),
        ("Stamp Duty (Cr)", params["stamp_duty"]),
        ("Sales/Marketing & OH (% of Revenue)", params["smoh_rate"]),
        ("Working Capital Interest Rate", params["interest_rate"]),
    ]

    ws_a["A1"] = "Refined Financial Model - Assumptions"
    ws_a["A1"].font = Font(bold=True, size=13)
    ws_a["A3"] = "Variable"
    ws_a["B3"] = "Value"
    ws_a["C3"] = "Source"
    for cell in (ws_a["A3"], ws_a["B3"], ws_a["C3"]):
        cell.fill = header_fill
        cell.font = header_font

    for i, (k, v) in enumerate(assumptions, start=4):
        ws_a.cell(i, 1, k)
        ws_a.cell(i, 2, v)
        ws_a.cell(i, 3, "Imported from original model (where available)")

    ws_a["A21"] = "Sales Mix by Year"
    ws_a["A21"].font = Font(bold=True)
    for i, v in enumerate(params["sales_mix"], start=22):
        ws_a.cell(i, 1, f"Y{i-21}")
        ws_a.cell(i, 2, v)

    ws_a["D21"] = "Construction Mix by Year"
    ws_a["D21"].font = Font(bold=True)
    for i, v in enumerate(params["cons_mix"], start=22):
        ws_a.cell(i, 4, f"Y{i-21}")
        ws_a.cell(i, 5, v)

    cols = [
        "Year",
        "Sales Mix",
        "Resi Price",
        "Comm Price",
        "Gross Revenue (Cr)",
        "LO Share Payout",
        "Net Revenue",
        "Construction Outflow",
        "Approval Outflow",
        "SM&OH",
        "Stamp Duty",
        "Deposit Out",
        "Deposit Refund",
        "Pre-Interest CF",
        "Opening Balance",
        "Interest",
        "Net CF",
        "Closing Balance",
    ]
    ws_m.append(cols)
    for c in range(1, len(cols) + 1):
        ws_m.cell(1, c).fill = header_fill
        ws_m.cell(1, c).font = header_font

    for r_idx, row in enumerate(base["rows"], start=2):
        ws_m.append(
            [
                row["year"],
                row["sales_mix"],
                row["resi_price"],
                row["comm_price"],
                row["gross_rev"],
                row["lo_payout"],
                row["net_rev"],
                row["cons_out"],
                row["appr_out"],
                row["smoh_out"],
                row["stamp"],
                row["dep_out"],
                row["dep_ref"],
                row["pre_interest_cf"],
                row["opening"],
                row["interest"],
                row["net_cf"],
                row["closing"],
            ]
        )

    ws_m.append([])
    ws_m.append(["KPI", "Value"])
    ws_m[7][0].fill = header_fill
    ws_m[7][0].font = header_font
    ws_m[7][1].fill = header_fill
    ws_m[7][1].font = header_font

    kpis = [
        ("Total Gross Revenue", base["kpi"]["gross_revenue"]),
        ("Total Cost (incl. interest, excl. LO share)", base["kpi"]["total_cost"]),
        ("Developer Net Cash Flow", base["kpi"]["net_cashflow"]),
        ("NPV @ 15%", base["kpi"]["npv_15"]),
        ("Model IRR", base["kpi"]["irr"]),
        ("Equity Multiple", base["kpi"]["equity_multiple"]),
    ]
    for i, (k, v) in enumerate(kpis, start=8):
        ws_m.cell(i, 1, k)
        ws_m.cell(i, 2, v)

    ws_s.append(["Scenario", "Total Revenue", "Developer Net CF", "NPV @15%", "IRR"])
    for c in range(1, 6):
        ws_s.cell(1, c).fill = header_fill
        ws_s.cell(1, c).font = header_font

    for row in scenario_rows:
        ws_s.append(row)

    ws_d.append(["Diagnostic Check", "Value", "Comment"])
    for c in range(1, 4):
        ws_d.cell(1, c).fill = header_fill
        ws_d.cell(1, c).font = header_font

    diagnostics = [
        ("Sales Mix Sum", sum(params["sales_mix"]), "Should be 1.0"),
        ("Construction Mix Sum", sum(params["cons_mix"]), "Should be 1.0"),
        ("Modeled Years", len(base["rows"]), "Aligned to project period assumption"),
        ("Original Model Project NPV Method", "Inflows-only in source", "Concerning: source excludes project outflows in row C8"),
        ("Original Interest Logic", "Hardcoded 32 Cr in source", "Concerning: source lacks debt drawdown schedule"),
    ]
    for d in diagnostics:
        ws_d.append(list(d))

    for ws in (ws_a, ws_m, ws_s, ws_d):
        for col in ws.columns:
            max_len = 0
            col_idx = col[0].column
            for cell in col:
                txt = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(txt), 60))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 2

    # number formatting
    for r in range(4, 30):
        if ws_a.cell(r, 2).value is not None and isinstance(ws_a.cell(r, 2).value, (int, float)):
            ws_a.cell(r, 2).number_format = "0.00"

    for r in range(2, 6):
        for c in range(2, 19):
            if isinstance(ws_m.cell(r, c).value, (int, float)):
                ws_m.cell(r, c).number_format = "0.00"

    for r in range(8, 20):
        if isinstance(ws_m.cell(r, 2).value, (int, float)):
            ws_m.cell(r, 2).number_format = "0.00"

    wb_new.save(REFINED_XLSX)

    # Build PDF summary
    styles = getSampleStyleSheet()
    story = []

    title = "ANA Feasibility Model - Investment Summary & Refinement"
    story.append(Paragraph(title, styles["Title"]))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%d %b %Y, %H:%M')}.", styles["Normal"]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("1) Variables Used", styles["Heading2"]))
    vars_table = Table(
        [
            ["Variable", "Value"],
            ["Land Area", f"{params['land_area_sqm']:,.2f} SqM"],
            ["Residential Area", f"{params['resi_area']:,.0f} Sft"],
            ["Commercial Area", f"{params['comm_area']:,.0f} Sft"],
            ["Start Price - Residential", f"Rs {params['price_resi']:,.0f}/sft"],
            ["Start Price - Commercial", f"Rs {params['price_comm']:,.0f}/sft"],
            ["Sales Growth", fmt_pct(params['sales_growth'])],
            ["Landowner Revenue Share", fmt_pct(params['lo_share'])],
            ["Construction Cost", fmt_cr(params['construction'])],
            ["Approvals & Liaisoning", fmt_cr(params['approval'])],
        ],
        hAlign="LEFT",
    )
    vars_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )
    story.append(vars_table)
    story.append(Spacer(1, 10))

    story.append(Paragraph("2) Concerning Aspects in Current Model", styles["Heading2"]))
    concerns = [
        "Project NPV in source workbook is computed from inflows-only (Cash Flow!C8), excluding project outflows; this overstates value.",
        "Interest cost is hardcoded at 32 Cr (ANA Summary row 35) and not linked to debt drawdown or timing.",
        "Key regulatory charges embed hardcoded reckoner rates and quantities in approval rows, reducing flexibility for updates.",
        "Parking and club-house revenue are currently zeroed in the base model, potentially understating revenue or indicating missing assumptions.",
        "Developer/Landowner sections have ambiguous labels (e.g., duplicate NPV labels), increasing interpretation risk.",
        "No explicit sensitivity table for price-growth / cost stress in the source workbook, limiting decision confidence.",
    ]
    for c in concerns:
        story.append(Paragraph(f"- {c}", styles["Normal"]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("3) Refined Base-Case Outputs", styles["Heading2"]))
    kpi_data = [
        ["Metric", "Output"],
        ["Total Gross Revenue", fmt_cr(base["kpi"]["gross_revenue"])],
        ["Total Cost (incl. interest, excl. LO share)", fmt_cr(base["kpi"]["total_cost"])],
        ["Developer Net Cash Flow", fmt_cr(base["kpi"]["net_cashflow"])],
        ["NPV @ 15%", fmt_cr(base["kpi"]["npv_15"])],
        ["Model IRR", fmt_pct(base["kpi"]["irr"])],
        ["Equity Multiple", f"{base['kpi']['equity_multiple']:.2f}x" if base['kpi']['equity_multiple'] is not None else "NA"],
    ]
    kpi_table = Table(kpi_data, hAlign="LEFT")
    kpi_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )
    story.append(kpi_table)
    story.append(Spacer(1, 10))

    story.append(Paragraph("4) Predictions (Scenario View)", styles["Heading2"]))
    scenario_table = Table([ ["Scenario", "Total Revenue", "Developer Net CF", "NPV @15%", "IRR"] ] + scenario_rows, hAlign="LEFT")
    scenario_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )
    story.append(scenario_table)
    story.append(Spacer(1, 10))

    story.append(Paragraph("5) Refined Working Files", styles["Heading2"]))
    story.append(Paragraph(f"- Refined workbook: {REFINED_XLSX.name}", styles["Normal"]))
    story.append(Paragraph(f"- Investment summary PDF: {SUMMARY_PDF.name}", styles["Normal"]))

    doc = SimpleDocTemplate(str(SUMMARY_PDF), pagesize=A4)
    doc.build(story)

    print(f"Generated: {REFINED_XLSX}")
    print(f"Generated: {SUMMARY_PDF}")


if __name__ == "__main__":
    main()
